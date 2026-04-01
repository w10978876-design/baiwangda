"""
快消品销售数据处理与标签热度分析脚本
=====================================
功能：
  1. 读取两渠道 Excel/TSV 数据，字段对齐、合并
  2. 地域解析（省份 / 城市 / 渠道类型）
  3. 商品名称解析（品牌 / 商品主体 / 口味 / 规格 / 包装）
  4. 统一大分类映射
  5. 标签热度计算（口味 / 商品主体 / 规格 / 价格带）
     热度 = 当前量级×40% + 时间趋势×35% + 地域广度×25%
  6. 输出 Excel 报告（标签热度总览 + 商品宽表）

使用方法：
  1. 修改下方 ══ 配置区 ══ 里的文件路径
  2. 运行：python fmcg_analysis.py
  3. 查看输出文件

依赖：pip install pandas openpyxl numpy
"""

import os
import re
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


# ══════════════════════════════════════════════════
# 配置区 — 根据实际情况修改这里
# ══════════════════════════════════════════════════

# 输入文件（支持 .xlsx / .xls / .txt / .tsv / .csv）
# 注意：必须与 CAT_MAP 的键一致——「渠道A」对应壹度表的大分类命名（需 CHANNEL_A_RENAME），
# 「渠道B」对应安达表的大分类命名。若两路文件与脚本标注对调，映射大量失效，页面会像「两套分类」重复出现。
INPUT_FILES = {
    "渠道A": "data/壹度2512-2601-2603-2503-2504.xlsx",
    "渠道B": "data/安达2512-2601-2603-2503-2504.xlsx",
}

# 渠道A的字段重命名（渠道A字段名与标准不同）
CHANNEL_A_RENAME = {
    "时间期间": "数据期间",
    "商品状态": "生命周期",
}

# 输出文件路径（固定写在脚本所在目录，避免「在别的目录运行」时 JSON 与 HTML 不同步导致页面 404）
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


def _abs_project_path(rel: str) -> str:
    rel = str(rel).strip()
    if os.path.isabs(rel):
        return rel
    return os.path.normpath(os.path.join(_SCRIPT_DIR, rel))


OUTPUT_FILE = _abs_project_path("商品标签热度分析.xlsx")

# 文件读取分隔符（TSV用'\t'，CSV用','，Excel文件改成 None）
FILE_SEP = None


# ══════════════════════════════════════════════════
# 词典区 — 根据业务补充和调整
# ══════════════════════════════════════════════════

# 统一大分类映射：(渠道来源, 原始大分类名) → 统一大分类
CAT_MAP = {
    ("渠道A", "FF类"):          "FF速食",
    ("渠道A", "休闲食品"):      "休闲零食",
    ("渠道A", "槟榔"):          "休闲零食",
    ("渠道A", "糖果巧克力"):    "休闲零食",
    ("渠道A", "饼干膨化"):      "休闲零食",
    ("渠道A", "低温日配"):      "日配冷藏",
    ("渠道A", "冲调方便速食"):  "方便速食",
    ("渠道A", "冷冻食品"):      "冷冻食品",
    ("渠道A", "常温奶"):        "常温乳品",
    ("渠道A", "常温饮料"):      "饮料",
    ("渠道A", "烘焙面包"):      "日配烘焙",
    ("渠道A", "粮油调味"):      "粮油调味",
    ("渠道A", "酒类"):          "酒类",
    ("渠道A", "日化美护"):      "日化美护",
    ("渠道A", "计生医疗"):      "日化美护",
    ("渠道A", "家居百货"):      "家居百货",

    ("渠道B", "加热速食"):      "FF速食",
    ("渠道B", "休闲零食"):      "休闲零食",
    ("渠道B", "饼干糖巧"):      "休闲零食",
    ("渠道B", "低温速食"):      "日配冷藏",
    ("渠道B", "日配冷藏"):      "日配冷藏",
    ("渠道B", "方便食品"):      "方便速食",
    ("渠道B", "冷冻"):          "冷冻食品",
    ("渠道B", "常温乳品"):      "常温乳品",
    ("渠道B", "饮料"):          "饮料",
    ("渠道B", "日配烘焙"):      "日配烘焙",
    ("渠道B", "粮油副食"):      "粮油调味",
    ("渠道B", "酒"):            "酒类",
    ("渠道B", "日化美护"):      "日化美护",
    ("渠道B", "家居百货"):      "家居百货",
}

# 已知城市白名单（防止误识别，可继续添加）
KNOWN_CITIES = [
    "包头", "呼和浩特", "赤峰", "通辽", "鄂尔多斯",
    "中牟", "濮阳", "郑州", "开封", "洛阳", "安阳", "焦作",
    "新乡", "许昌", "漯河", "南阳", "商丘", "信阳", "周口", "驻马店",
    "北京", "上海", "广州", "深圳", "成都", "武汉", "西安",
    "杭州", "南京", "天津", "重庆", "沈阳", "长春", "哈尔滨",
    "济南", "青岛", "合肥", "福州", "厦门", "南昌",
    "长沙", "昆明", "贵阳", "南宁", "海口",
    "太原", "石家庄", "兰州", "西宁", "银川", "乌鲁木齐",
]

# 渠道类型关键词映射
CHANNEL_MAP = {
    "便利":   "便利店",
    "超市":   "超市",
    "大卖场": "大卖场",
    "社区":   "社区店",
    "线上":   "线上",
    "电商":   "线上",
    "KA":     "KA卖场",
    "分公司": "分公司门店",
}

# 省份列表
PROVINCES = [
    "北京", "天津", "上海", "重庆",
    "河北", "山西", "辽宁", "吉林", "黑龙江",
    "江苏", "浙江", "安徽", "福建", "江西", "山东",
    "河南", "湖北", "湖南", "广东", "海南",
    "四川", "贵州", "云南", "陕西", "甘肃", "青海",
    "内蒙古", "广西", "西藏", "宁夏", "新疆",
]

# 口味词典（解析时按字符串长度从长到短匹配，复合/长尾词请写完整如「青柠味」）
FLAVOR_WORDS = [
    "日式浇汁", "蜂蜜芥末", "火鸡面风味", "黑胡椒", "奥尔良",
    "照烧", "香辣", "麻辣", "红烧", "咖喱", "蒜香", "酱香",
    "香菇", "糯米", "蟹味", "原味", "清淡", "甜味", "咸味",
    "番茄", "草莓", "芒果", "抹茶", "黑糖", "海鲜", "芝士",
    "浸汁", "炭烤",
    "青柠味", "柠檬味", "白桃味", "水蜜桃味", "葡萄味", "柑橘味", "西柚味",
    "葡萄伏特加风味", "热带风味", "血橙味",
    "荔枝海盐味", "海盐荔枝味", "荔枝味",
    "水蜜桃", "柠檬", "青柠", "白桃", "荔枝",
    # 高频遗漏（跨品类）
    "蜜桃味", "鸡翅味", "微辣", "香辣味", "麻辣味", "五香味",
    "卤味", "孜然味", "牛肉味", "鸡肉味", "猪肉味",
    "果味", "苹果味", "桃味", "梨味", "菠萝味",
    "酸梅味", "杨梅味", "玫瑰味", "茉莉味",
    "枇杷炖梨", "青提茉莉风味", "无糖", "0糖", "椰蓉",
    "巧克力味", "巧克力",
    "香草味", "西梅", "芝麻味",
    "炭烧", "红心芭乐味", "金汤肥牛", "老母鸡汤",
    "糖醋味", "蜂蜜黄油味", "麻酱味", "爆辣味", "韩国泡菜味", "正梅味",     "泡椒", "山椒",
    "重庆火锅味", "川香",
    "湘辣", "黑椒", "酸菜牛肉", "老坛酸菜牛肉", "加辣",
    "虾味", "黄瓜味", "蓝莓味", "牛乳味", "红豆",
    "草莓桑葚味", "草莓味", "桑葚味", "牛油果", "水果味",
    "燕麦+黄桃", "牛奶蛋羹",
    "青苹果", "红苹果", "苹果",
    # ↑ 在这里继续补充遗漏的口味词
]

FLAVOR_WORDS_SORTED = sorted(set(FLAVOR_WORDS), key=len, reverse=True)

# 统计/热度聚合用：合并「苹果味」与「苹果」「红苹果」等同义写法，避免同一口味拆成多条信号
_FLAVOR_STATS_KEEP_WEI = frozenset(
    {"原味", "无味", "寡味", "五味", "独味", "怪味", "口味"}
)
# 显式别名（在去「味/风味」后缀之后再应用一次；无别名则仅靠去「味」即可合并如 苹果味→苹果）
FLAVOR_STATS_ALIASES = {
    "红苹果": "苹果",
    "青苹果": "苹果",
    "黄苹果": "苹果",
}
# 蜜桃/水蜜桃/白桃等统计口径统一为「桃」（樱桃/杨桃等不含下列键，避免误伤）
FLAVOR_STATS_PEACH_TO_TAO = {
    "蜜桃": "桃",
    "水蜜桃": "桃",
    "白桃": "桃",
    "黄桃": "桃",
    "脆桃": "桃",
    "冬桃": "桃",
    "雪桃": "桃",
    "桃子": "桃",
    "蜜桃味": "桃",
    "水蜜桃味": "桃",
    "白桃味": "桃",
    "黄桃味": "桃",
    "桃味": "桃",
    "桃子味": "桃",
}

# 商品主体：热度聚合用（明细列仍可保留酸牛奶/奶酪益生元等写法）
BODY_STATS_ALIASES = {
    "酸牛奶": "酸奶",
    "奶酪益生元": "奶酪",
    "精制食用盐": "食盐",
    "食用盐": "食盐",
    "葱香手抓饼": "手抓饼",
    # 覆盖表/历史主体整段写法，热度聚合时与「鸡排」合并
    "黑椒鸡排": "鸡排",
    "奥尔良鸡排": "鸡排",
    "香辣鸡排": "鸡排",
    # 方便面系列名归一为「方便面」，形态在包装类型列体现
    "方便面一袋半": "方便面",
    "方便面一倍半": "方便面",
    "臻享浓牛奶": "牛奶",
    "臻浓牛奶": "牛奶",
    "果粒优酸乳黄": "优酸乳",
    "蓝莓优酸乳": "优酸乳",
    "牛乳口味脆筒": "冰淇淋脆筒",
    "冰淇淋巧脆棒": "冰淇淋巧脆棒",
    "京门爆肚麻辣": "素毛肚",
    "湘辣小公鱼": "小鱼干",
    "咪咪虾条虾味": "虾条",
    "乐事黄瓜味": "薯片",
    "黑椒北道烧": "北道烧",
    "自煮火锅": "自热火锅",
    # 品牌误入主体兜底
    "科迪": "纯牛奶",
    # 日配冷藏中该主体更准确命名
    "酪蛋白": "酪蛋白牛奶",
}


def normalize_body_for_stats(body) -> str:
    if body is None or (isinstance(body, float) and pd.isna(body)):
        return ""
    s = str(body).strip()
    if not s or s.lower() == "nan":
        return ""
    s = BODY_STATS_ALIASES.get(s, s)
    # 火锅场景：自煮/自热统一口径
    if s in ("自煮火锅", "自热火锅", "自热食品", "自煮食品"):
        return "自热火锅"
    # 通用冰淇淋形态归一
    if "脆筒" in s:
        return "冰淇淋脆筒"
    if "巧脆棒" in s:
        return "冰淇淋巧脆棒"
    if "方糕" in s:
        return "冰淇淋方糕"
    if "冰棍" in s:
        return "冰棍"
    # 通用零食形态归一
    if "虾条" in s:
        return "虾条"
    if "薯片" in s:
        return "薯片"
    if "爆肚" in s:
        return "素毛肚"
    if "小公鱼" in s or "小鱼" in s:
        return "小鱼干"
    # 乳品常见混写归一
    if "优酸乳" in s:
        return "优酸乳"
    if s in ("臻享浓牛奶", "臻浓牛奶", "金典有机奶"):
        return "牛奶"
    if "手抓饼" in s and s != "手抓饼":
        return "手抓饼"
    return s


def normalize_flavor_for_stats(flavor) -> str:
    """将解析得到的口味文本规范为统计口径（仅用于热度聚合与报表标签值，不改写明细解析列）。"""
    if flavor is None or (isinstance(flavor, float) and pd.isna(flavor)):
        return ""
    s = str(flavor).strip()
    if not s or s.lower() == "nan":
        return ""
    # 业务要求：保留「果味」原词，不做「果」或「水果」归并
    if s == "果味":
        return "果味"

    def _apply_aliases(x: str) -> str:
        return FLAVOR_STATS_ALIASES.get(x, x)

    s = _apply_aliases(s)
    changed = True
    while changed and s:
        changed = False
        if s.endswith("风味") and len(s) > len("风味"):
            s = s[: -len("风味")].strip()
            s = _apply_aliases(s)
            changed = True
            continue
        if s.endswith("味") and len(s) > 1 and s not in _FLAVOR_STATS_KEEP_WEI:
            s = s[:-1].strip()
            s = _apply_aliases(s)
            changed = True
    return FLAVOR_STATS_PEACH_TO_TAO.get(s, s)


# 饮料等：名称末尾为标准品类词时，后缀优先于「中间截断」匹配主体（如 枇杷炖梨+水果饮料）
# (后缀子串, 标准主体标签)，运行时按后缀长度从长到短尝试
# 乳品/低温：按后缀从长到短锁定主体，余量再抽口味（避免「畅轻燕麦+黄桃」整段当主体）
_DAIRY_BODY_SUFFIXES = [
    ("风味发酵乳", "发酵乳"),
    ("发酵乳", "发酵乳"),
    ("酸牛奶", "酸牛奶"),
    ("椰子水", "椰子水"),
    ("轻食杯", "轻食杯"),
    ("酸奶", "酸奶"),
    ("乳酸菌", "酸奶"),
    ("牛奶", "牛奶"),
]

END_BODY_SUFFIX_BY_CAT = {
    "饮料": [
        ("水果饮料", "水果饮料"),
        ("红苹果汁", "红苹果汁"),
        ("苹果汁", "苹果汁"),
        ("果蔬汁", "果蔬汁"),
        ("果汁饮料", "果汁饮料"),
        ("植物饮料", "植物饮料"),
        ("蛋白饮料", "蛋白饮料"),
        ("乳酸菌饮料", "乳酸菌饮料"),
        ("茶饮料", "茶饮料"),
        ("碳酸饮料", "碳酸饮料"),
        ("运动饮料", "运动饮料"),
        ("能量饮料", "能量饮料"),
        ("维生素饮料", "维生素饮料"),
        ("可乐", "可乐"),
        ("矿泉水", "矿泉水"),
        ("纯净水", "纯净水"),
        ("苏打水", "苏打水"),
        ("气泡水", "气泡水"),
        ("饮用水", "饮用水"),
    ],
    "日配冷藏": list(_DAIRY_BODY_SUFFIXES),
    "常温乳品": list(_DAIRY_BODY_SUFFIXES),
}

# 商品名开头的已知品牌/子品牌（越长越优先匹配，用于无字段或字段明显错误时纠偏）
KNOWN_BRAND_PREFIXES = sorted(
    {
        "农夫山泉东方树叶",
        "农夫山泉水溶C",
        "元气森林外星人",
        "元气森林",
        "农夫山泉",
        "东方树叶",
        "水溶C",
        "江小白果立方",
        "江小白",
        "凤城老窖",
        "海底捞",
        "双汇",
        "壹度定制",
        "五得利",
        "红星",
        "汾酒",
        "郎酒",
        "锐澳微醺",
        "锐澳",
        "青岛啤酒",
        "娃哈哈",
        "康师傅",
        "统一企业",
        "统一",
        "怡宝至本清润",
        "怡宝",
        "伊利畅轻",
        "趣味多",
        "每日博士",
        "妙可蓝多",
        "悦鲜活",
        "牧民人家",
        "洁婷",
        "苏菲",
        "淘淘氧棉",
        "谷淦",
        "杜蕾斯",
        "冈本",
        "高洁丝",
        "幸运",
        "杨掌柜",
        "王中王",
        "大辣娇",
        "一桶半",
        "金典",
        "旺仔",
        "养元",
        "六个核桃",
        "惠丫丫",
        "卡塔",
        "卡拉美拉",
        "新海",
        "壹汀定制",
        "倩丽",
        "三全",
        "思念",
        "树熟",
        "巧乐兹",
        "可莎",
        "盈泰",
        "劲仔",
        "呀土豆",
        "张新发",
        "竹香园",
        "都市牧场",
        "怡冠园",
        "好友趣",
        "葛小宝",
        "印迹",
        "鼎味泰",
        "锦甜",
        "田小花",
        "友厨坊",
        "金厨娘",
        "佳龙",
        "飞旺",
        "特仑苏",
        "臻享浓",
        "臻浓",
        "科迪",
        "好欢螺",
        "礼拜天",
        "未来星",
        "咪咪",
        "乐事",
        "新希望荷荷",
        "味全好喝椰",
        "壹度便利",
        "百岁山",
        "名仁",
        "补水啦",
        "百事可乐零度",
        "可口可乐",
        "百事",
        "雪碧",
        "芬达",
        "东鹏",
        "红牛",
        "脉动",
        "外星人",
        # 粮油调味
        "王守义",
        "加加",
        "海天",
        "李锦记",
        "老干妈",
        "太太乐",
        "味好美",
        "厨邦",
        "欣和",
        "六月鲜",
        # 零食
        "卫龙",
        "三只松鼠",
        "良品铺子",
        "百草味",
        # 肉制品（双汇已上）
        "金锣",
        "雨润",
        "郑荣",
        # 乳品
        "蒙牛",
        "伊利",
        "味全",
        "海河",
        "光明",
        "新希望",
        "君乐宝简醇",
        "君乐宝",
        "安琪",
        # 方便速食
        "今麦郎",
        "白象",
        # 渠道/区域品牌与系列
        "安哒",
        "川南",
        "梅见",
        "诺贝达",
        "桃李",
        "新欧乐",
    },
    key=len,
    reverse=True,
)

# ERP「商品品牌」明显占位/错误时，以商品名称前缀推断为准
BRAND_FIELD_BLACKLIST = frozenset({"圣诞", "元旦", "测试", "未知", "无", "—", "-"})

# 包装类型词典
PACKAGE_WORDS = [
    "单包装", "挂式抽纸", "原生枕", "梦幻盖", "利乐枕", "利乐包", "利乐钻", "钻包",
    "利乐苗条装",
    "大食桶", "大食袋", "一倍半", "一袋半",
    "罐装", "瓶装", "袋装", "盒装", "礼盒", "散装", "桶装", "箱装",
]

# 从主体串中剔除的常见包装/渠道后缀（避免「牛奶利乐梦幻盖」类主体）
# 注意：大食桶/一袋半等若作为「方便面」形态的一部分写在主体里，不要在此剔除（见 refine 方便面大食桶）
_BODY_PACKAGING_STRIP = [
    "原生枕", "梦幻盖", "利乐枕", "利乐包", "利乐钻", "钻包", "钻石包",
    "利乐苗条装",
    "随心配", "福字",
]


def _strip_body_packaging_tokens(body: str) -> str:
    b = str(body).strip() if body else ""
    if not b:
        return b
    orig = b
    for tok in sorted(_BODY_PACKAGING_STRIP, key=len, reverse=True):
        b = b.replace(tok, "")
    b = re.sub(r"[A-Za-z]{1,4}$", "", b)
    b = re.sub(r"^[\s\*]+|[\s\*]+$", "", b)
    b = re.sub(r"\*+\d*$", "", b)
    b = b.strip()
    return b if b else orig


def _strip_brand_from_body(body: str, brand: str) -> str:
    """主体误带品牌前缀时剥离，如「德芙脆香米」→「脆香米」."""
    b = str(body).strip() if body else ""
    br = str(brand).strip() if brand else ""
    if not b or not br:
        return b
    if b.startswith(br) and len(b) > len(br):
        tail = b[len(br):].strip()
        if tail:
            return tail
    return b

# 商品主体词典（具体词放前面优先匹配，可继续补充）
PRODUCT_TYPES = [
    # FF速食 ─────────────────
    (r"鸡肉包",              "鸡肉包"),
    (r"牛肉包",              "牛肉包"),
    (r"油菜包",              "油菜包"),
    (r"肉包",                "肉包"),
    (r"烧麦",                "烧麦"),
    (r"鸡肉饭团|饭团",       "鸡肉饭团"),
    (r"地道肠",              "地道肠"),
    (r"鸡肉肠|开花鸡肉肠",  "鸡肉肠"),
    (r"鸡肉串",              "鸡肉串"),
    (r"腿肉串",              "腿肉串"),
    (r"猪肉串",              "猪肉串"),
    (r"唐扬串",              "唐扬串"),
    (r"炭烤鸡全腿|鸡全腿",  "鸡全腿"),
    (r"鸡丸串|丸串",         "鸡丸串"),
    (r"海带串",              "海带串"),
    (r"面筋",                "面筋"),
    (r"卤蛋",                "卤蛋"),
    (r"蛋挞",                "蛋挞"),
    (r"熔岩面包|天然酵母面包|酵母面包|菠萝包", "面包"),
    (r"玉米",                "玉米"),
    (r"豆奶|豆浆",           "豆奶"),
    # 酒类 ───────────────────
    (r"二锅头",              "二锅头"),
    (r"高粱酒",              "高粱酒"),
    (r"白酒",                "白酒"),
    (r"啤酒",                "啤酒"),
    (r"葡萄酒|红酒",         "葡萄酒"),
    (r"鸡尾酒|强爽",          "鸡尾酒"),
    (r"果汁酒",               "果汁酒"),
    (r"地瓜烧",               "地瓜烧"),
    (r"茉莉花茶",             "茉莉花茶"),
    (r"冰红茶饮料|冰红茶",     "冰红茶饮料"),
    (r"复合果汁",             "复合果汁"),
    (r"午餐肉罐头|午餐肉",     "午餐肉罐头"),
    (r"白砂糖",               "白砂糖"),
    (r"小麦粉",               "小麦粉"),
    (r"火锅底料",             "火锅底料"),
    # 饮料水饮（具体词放前，避免「电解质」早于「电解质水」被截断）────────
    (r"电解质水|电解制水",    "电解质水"),
    (r"^电解质$|^电解$",      "电解质水"),
    (r"苏打水|苏打气泡水",    "苏打水"),
    (r"纯净水|蒸馏水|纯水",    "纯净水"),
    (r"天然矿泉水|矿泉水|天然水", "矿泉水"),
    (r"气泡水|含气水",        "气泡水"),
    (r"饮用水|直饮水",        "饮用水"),
    # ↑ 在这里继续补充遗漏的主体词（具体词放前面）
]

# 品类感知主体词：先匹配当前「统一大分类」下的规则，再回退 PRODUCT_TYPES（全品类共用）
# 键必须与 CAT_MAP 产出的统一大分类名称一致。
CAT_PRODUCT_TYPES = {
    "粮油调味": [
        (r"麻辣鲜", "麻辣鲜"),
        (r"十三香", "十三香"),
        (r"味精|蔬之鲜", "味精"),
        (r"鸡精|鸡粉", "鸡精"),
        (r"榨菜|咸菜|酸菜|泡菜|橄榄菜", "榨菜"),
        (r"生抽|老抽|酱油|味极鲜|鲜酱油", "酱油"),
        (r"陈醋|米醋|香醋|白醋|黑醋", "醋"),
        (r"料酒|烹调黄酒|烹饪黄酒", "料酒"),
        (r"蚝油|鲍鱼汁", "蚝油"),
        (r"豆瓣酱|黄豆酱|甜面酱|辣椒酱|拌饭酱|牛肉酱|香菇酱", "调味酱"),
        (r"腐乳|豆豉|霉豆腐", "腐乳"),
        (r"芝麻酱|花生酱|番茄酱|沙拉酱", "酱类"),
        (r"花椒|八角|桂皮|香叶|孜然|辣椒粉", "香辛料"),
        (r"大米|粳米|籼米|糯米|香米|小米|杂粮", "米类"),
        # 将“粉面原料”细分，避免主体过泛（如 面粉/玉米淀粉/红薯淀粉 都应落到具体项）
        (r"红薯淀粉", "红薯淀粉"),
        (r"玉米淀粉", "玉米淀粉"),
        (r"面粉|小麦粉", "面粉"),
        (r"生粉", "生粉"),
        (r"挂面|龙须面|刀削面|方便面饼", "面条"),
        (r"粉丝|粉条|宽粉|米线", "粉类"),
        (r"食用油|花生油|大豆油|菜籽油|玉米油|橄榄油|调和油", "食用油"),
        (r"精制食用盐|精制盐|食盐|食用盐|海盐|低钠盐|加碘盐", "食盐"),
        (r"冰糖|红糖|绵白糖", "食糖"),
        (r"下饭菜酱腌菜|酱腌菜|下饭菜", "下饭菜酱腌菜"),
        (r"安琪酵母|高活性干酵母|干酵母|酵母粉|鲜酵母|耐高糖酵母|^酵母$", "酵母粉"),
    ],
    "饮料": [
        (r"维生素饮料|维他命水|维\d*饮料", "维生素饮料"),
        (r"运动饮料|能量饮料|功能饮料", "功能饮料"),
        (r"茶饮料|奶茶|乳茶", "茶饮料"),
        (r"果蔬汁|果汁饮料|果味饮料", "果汁饮料"),
        (r"水果饮料|植物饮料", "水果饮料"),
    ],
    "休闲零食": [
        (r"京门爆肚|素爆肚|爆肚", "素毛肚"),
        (r"酸奶棒糖", "酸奶棒糖"),
        (r"酸奶奶棒", "酸奶奶棒"),
        (r"臭干子|臭豆腐干", "豆干零食"),
        (r"小公鱼", "小鱼干"),
        (r"海春笋|山椒笋|泡椒笋", "笋类零食"),
        (r"老卤鸭掌|卤鸭掌|麻辣鸭掌|鸭掌", "鸭掌"),
        (r"口嚼葛根", "口嚼葛根"),
        (r"好友趣", "薯片"),
        (r"都市牧场", "含片"),
        (r"含片糖|薄荷糖", "含片糖"),
        (r"素牛排|牛排", "素牛排"),
        (r"好友趣|呀土豆", "膨化食品"),
        (r"小鱼", "小鱼干"),
        (r"脆笋", "笋类零食"),
        (r"槟榔", "槟榔"),
        (r"泡泡糖", "泡泡糖"),
        (r"葛根", "葛根零食"),
        (r"迷你牛角包|牛角包", "牛角包"),
        (r"微笑鸡蛋仔", "微笑鸡蛋仔"),
        (r"鸡蛋仔", "鸡蛋仔"),
        (r"扭扭棒", "扭扭棒"),
        (r"满格华夫", "满格华夫饼"),
        (r"华夫饼|华夫", "华夫饼"),
        (r"混糖月饼|月饼", "月饼"),
        (r"干脆面|干吃面|点心面|干吃(?![面])", "干脆面"),
        (r"辣条|辣片|面筋制品|魔芋爽|素肉", "辣条"),
        (r"薯片|薯条|山药片|锅巴|米饼", "膨化食品"),
        (r"瓜子|花生|坚果|巴旦木|腰果|核桃", "坚果炒货"),
        (r"肉干|肉脯|凤爪|鸭脖|卤味零食", "肉干卤味"),
        (r"饼干|曲奇|威化|苏打饼干", "饼干"),
        (r"巧克力(?![味菠])|糖果|软糖|硬糖|口香糖", "糖巧"),
    ],
    "方便速食": [
        (r"螺蛳粉", "螺蛳粉"),
        (r"粉面菜蛋", "粉面菜蛋"),
        (r"毛血旺|川香毛血旺|火锅川香|川香.*毛血旺", "自热火锅"),
        (r"火腿肠|王中王", "火腿肠"),
        (r"拌面|汤面|排骨面|酸菜牛肉面|牛肉面|葱香排骨面|酸辣牛肉面", "方便面"),
        # 勿把「大食桶/一袋半」等单写进本组，否则会只匹配到桶/袋本身，破坏口味与 refine 组合
        (r"方便面|泡面|面饼|桶面|袋面", "方便面"),
        (r"自热米饭|自热火锅|自热食品", "自热食品"),
        (r"速食粥|速食汤|冲泡粉丝", "冲泡速食"),
    ],
    "冷冻食品": [
        (r"香草口味脆筒|香草味脆筒|香草脆筒|脆筒", "冰淇淋脆筒"),
        (r"巧脆棒|脆棒", "冰淇淋巧脆棒"),
        (r"墨鱼爆蛋|章鱼爆蛋", "火锅丸滑"),
        (r"黄金包馅鱿鱼卷|鱿鱼卷|黄金脆骨棒|脆骨棒|墨鱼丸", "火锅丸滑"),
        (r"北道烧|黑椒北道烧", "北道烧"),
        (r"冰淇淋|冰激凌|雪糕", "冰淇淋"),
        (r"冰杯|食用冰杯", "冰杯"),
        (r"榴莲肉", "榴莲肉"),
        (r"汤圆", "汤圆"),
        (r"冰淇淋|冰激凌|雪糕", "冰淇淋"),
        (r"速冻水饺|速冻馄饨|饺子|云吞", "速冻面点"),
        (r"速冻汤圆|元宵", "速冻甜品"),
        (r"火锅丸|撒尿牛丸|鱼丸|虾滑", "火锅丸滑"),
    ],
    "FF速食": [
        (r"葱香手抓饼|手抓饼", "手抓饼"),
        (r"黑椒鸡排|奥尔良鸡排|香辣鸡排|鸡排", "鸡排"),
        (r"地道大肉肠|大肉肠", "烤肠"),
        (r"台式香肠", "台式香肠"),
        (r"香肠", "香肠"),
        (r"饭团|寿司|便当", "饭团便当"),
        (r"烤肠|热狗|地道肠", "烤肠"),
    ],
    "常温乳品": [
        (r"AD钙奶", "AD钙奶"),
        (r"果粒优酸乳", "优酸乳"),
        (r"利乐枕早餐奶|早餐奶", "早餐奶"),
        (r"臻享浓牛奶|臻浓牛奶", "牛奶"),
        (r"特仑苏牛奶", "牛奶"),
        # 「特仑苏」已作品牌剥离后，余量常为「牛奶利乐…」；需仍能识别为牛奶
        (r"(?<!纯)(?<!酸)(?<!鲜)(?<!配方)牛奶(?!粉|酪)", "牛奶"),
        (r"奶茶", "奶茶"),
        (r"有机纯|纯奶|纯牛奶|有机纯奶|有机牛奶", "纯牛奶"),
        (r"六个核桃|核桃乳", "核桃乳"),
        (r"旺仔牛奶复原乳|复原乳", "牛奶"),
        (r"椰子水", "椰子水"),
        (r"轻食杯", "轻食杯"),
        (r"酸牛奶", "酸牛奶"),
        (r"发酵乳|风味发酵乳", "发酵乳"),
        (r"纯牛奶|鲜牛奶|低脂奶|脱脂奶|高钙奶", "液态奶"),
        (r"酸奶|乳酸菌", "酸奶"),
        (r"奶粉|成人奶粉", "奶粉"),
    ],
    "日配冷藏": [
        (r"北道烧|黑椒北道烧", "北道烧"),
        (r"椰子水", "椰子水"),
        (r"轻食杯", "轻食杯"),
        (r"酸牛奶", "酸牛奶"),
        (r"发酵乳|风味发酵乳", "发酵乳"),
        (r"酸奶|乳酸菌", "酸奶"),
        # 其它低温乳基/乳基饮品：避免 fallback 把整串当主体
        (r"纤维饮", "纤维饮"),
        (r"奶酪益生元|益生元奶酪", "奶酪益生元"),
        (r"高钙奶酪棒|奶酪棒", "奶酪棒"),
        (r"A2?β?酪蛋白|β?酪蛋白|酪蛋白", "酪蛋白牛奶"),
        (r"鲜奶|巴氏奶|低温奶", "鲜奶"),
        (r"冷藏酸奶|活菌", "冷藏酸奶"),
    ],
    "日配烘焙": [
        (r"迷你牛角包", "迷你牛角包"),
        (r"牛角包", "牛角包"),
        # 部分 SKU 写作「酵母（巧克力味）」省略「面包」，此处将“酵母”归为面包主体
        (r"酵母(?!粉)", "面包"),
        (r"青团", "青团"),
        (r"熔岩面包|天然酵母面包|酵母面包|菠萝包", "面包"),
        (r"起酥苹果|起酥面包|迷你起酥|起酥", "起酥面包"),
        (r"面包|吐司|餐包|欧包", "面包"),
        (r"蛋糕|瑞士卷|慕斯", "蛋糕"),
    ],
    "酒类": [
        (r"全麦白啤|白啤", "全麦白啤"),
        (r"梅酒|青梅酒", "梅酒"),
        (r"白酒|高粱酒|二锅头|玻汾|老白干", "白酒"),
        (r"啤酒|精酿|原浆", "啤酒"),
        (r"葡萄酒|红酒|干红|干白", "葡萄酒"),
        (r"洋酒|威士忌|伏特加|白兰地", "洋酒"),
        (r"果酒|预调酒|鸡尾酒|微醺", "鸡尾酒"),
    ],
    "日化美护": [
        (r"避孕套|安全套", "避孕套"),
        (r"洗发水|洗发露|护发素", "洗发护发"),
        (r"牙膏|牙刷|漱口水", "口腔护理"),
        (r"洗衣液|洗衣粉|洗衣凝珠|柔顺剂", "衣物清洁"),
        (r"面巾纸|餐巾纸|手帕纸|洗脸巾", "面巾纸"),
        (r"抽纸|纸巾|湿巾|卫生纸", "纸品"),
        # 个人护理：成人用品/护理巾等显式主体兜底
        (r"安心裤|拉拉裤", "安心裤"),
        (r"尿裤|纸尿裤|纸尿片", "尿裤"),
        (r"卫生巾|姨妈巾|护垫", "卫生巾"),
        (r"透气.*日夜用|纯棉.*日夜用", "卫生巾"),
        (r"超长夜用|日夜用|夜用|日用", "卫生巾"),
    ],
    "家居百货": [
        (r"打火机", "打火机"),
        (r"棉袜|船袜|袜", "袜子"),
        (r"晴雨两用伞|遮阳伞|雨伞|伞", "雨伞"),
        (r"雨披", "雨披"),
        (r"扑克", "扑克"),
        (r"塑杯", "塑杯"),
        (r"垃圾袋|保鲜膜|一次性手套", "家居耗材"),
        (r"电池|插座|灯泡", "小五金电"),
    ],
}

# 去掉口味/规格后剩余短词 → 规范主体（如「电解质」→「电解质水」）
BODY_REMAINDER_MAP = {
    "电解质": "电解质水",
    "电解": "电解质水",
    "纯水": "纯净水",
    "矿泉水": "矿泉水",
    "苏打": "苏打水",
}

# 已锁定标准主体后，若口味词表仍无命中，可用「主体→常见口味」补全（复合调味/零食名常用）
BODY_TO_FLAVOR_HINT = {
    "麻辣鲜": "麻辣",
    "辣条": "香辣",
}


def _normalize_product_title(s: str) -> str:
    """去掉渠道常见前缀（如【新】），避免干扰品牌/主体边界。"""
    t = str(s).strip()
    t = re.sub(r"^[【\[]\s*新\s*[】\]]\s*", "", t)
    t = re.sub(r"^[（(]\s*新\s*[）)]\s*", "", t)
    t = re.sub(r"\*+\d+\s*$", "", t).strip()
    return t.strip()


def _strip_leading_ascii_noise(s: str) -> str:
    """去掉 1～3 位英文前缀（如 z / D 渠道码），但保留 AD钙奶 等以 AD 开头的品名。"""
    t = re.sub(r"^[^一-龥A-Za-z0-9]+", "", str(s).strip())
    # 常见「DAD钙奶」类：去掉前导 D，保留 AD钙奶 品名词头
    t = re.sub(r"^D+(?=AD钙)", "", t)
    if re.match(r"^AD钙", t):
        return t
    return re.sub(r"^[A-Za-z]{1,3}\s*", "", t)


def _lock_body_suffix_first(work: str, unified_cat: str) -> tuple[str, str]:
    """按名称末尾品类词锁定主体（如 …枇杷炖梨水果饮料 → 水果饮料）。"""
    if not work or not str(work).strip():
        return "", ""
    cat = str(unified_cat).strip()
    items = END_BODY_SUFFIX_BY_CAT.get(cat, [])
    for suffix, label in sorted(items, key=lambda x: len(x[0]), reverse=True):
        if work.endswith(suffix):
            return label, suffix
    return "", ""


def _try_mooncake_sugar_prefix(work: str) -> tuple[str, str, str]:
    """糖醇/冰糖 + 混糖月饼 → 主体混糖月饼，口味为前缀糖类型。"""
    m = re.match(r"^(糖醇|冰糖|木糖醇)(混糖月饼)$", work)
    if m:
        return "混糖月饼", m.group(0), m.group(1)
    return "", "", ""


def _lock_body_label(work: str, unified_cat: str) -> tuple[str, str]:
    """
    在「去品牌、去度数」后的 work 上，按品类优先锁定标准主体。
    返回 (标准主体标签, work 中实际匹配到的子串)；无则 ("", "").
    先锁主体再抽口味，避免「麻辣」等子串拆掉「麻辣鲜」。
    """
    if not work or not str(work).strip():
        return "", ""
    cat = str(unified_cat).strip()
    for pattern, label in CAT_PRODUCT_TYPES.get(cat, []):
        m = re.search(pattern, work)
        if m:
            return label, m.group(0)
    for pattern, label in PRODUCT_TYPES:
        m = re.search(pattern, work)
        if m:
            return label, m.group(0)
    return "", ""


# 去品牌后若整段只剩这些字，多为产品线/广告字而非真实商品主体（截断名常见）
PRODUCT_LINE_ONLY_MARKERS = frozenset({
    "鲜", "香", "新", "浓", "醇", "纯", "爽", "轻", "真",
})
PRODUCT_LINE_ONLY_MARKERS_2 = frozenset({"经典", "金装", "优选", "特选"})


def _coerce_body_if_only_product_line(body: str, brand: str, raw_norm: str, unified_cat: str) -> str:
    """截断导致的「品牌+鲜/香…」单字尾巴，映射为合理主体，避免把「鲜」当成商品主体。"""
    b = str(body).strip()
    cat = str(unified_cat).strip()
    if not b:
        return body
    if b not in PRODUCT_LINE_ONLY_MARKERS and b not in PRODUCT_LINE_ONLY_MARKERS_2:
        return body
    if "十三香" in raw_norm:
        return "十三香"
    if cat == "粮油调味":
        return "调味料"
    if cat == "饮料":
        return "饮料"
    if cat in ("休闲零食", "方便速食", "FF速食"):
        return "休闲食品"
    if cat == "酒类":
        return "酒类"
    return "调味料"

# 规格正则（容量 / 重量，与度数严格区分）
SPEC_RE = re.compile(
    r"(\d+(?:\.\d+)?)\s*"
    r"(ml|ML|mL|L|升|g|G|kg|KG|克|斤|包|瓶|罐|袋|片|粒|支|盒|桶|箱|个|串|根|条|只)"
)

# 度数正则（酒类特有，归入口味字段）
DEGREE_RE = re.compile(r"(\d+(?:\.\d+)?)\s*[°度]")

# 价格带分箱
PRICE_BINS   = [0, 3, 6, 10, 20, 9999]
PRICE_LABELS = ["低价(≤3元)", "亲民(3-6元)", "中价(6-10元)", "中高(10-20元)", "高价(20元+)"]


# ══════════════════════════════════════════════════
# 解析函数
# ══════════════════════════════════════════════════

def read_file(path, source_name, sep="\t"):
    """读取数据文件，自动识别 Excel 与文本格式"""
    if sep is None or path.endswith((".xlsx", ".xls")):
        df = pd.read_excel(path)
    else:
        df = pd.read_csv(path, sep=sep)
    df["渠道来源"] = source_name
    return df


def parse_region(name):
    """
    解析地域名称 → 省份 / 城市 / 渠道类型

    格式1：'中牟-张志鹏'       → 城市=中牟（连字符前为城市）
    格式2：'安哒便利包头分公司' → 城市=包头（白名单扫描）
    """
    name = str(name).strip()
    province = city = channel = ""

    for p in PROVINCES:
        if p in name:
            province = p
            break

    m = re.match(r"^([^\-\—]+)[\-\—]", name)
    if m:
        city = m.group(1).strip()
    if not city or city not in KNOWN_CITIES:
        for c in KNOWN_CITIES:
            if c in name:
                city = c
                break

    for kw, ch in CHANNEL_MAP.items():
        if kw in name:
            channel = ch
            break

    return pd.Series([province, city, channel])


def _infer_brand_prefix(name: str) -> str:
    """取商品名最前缀的已知品牌（长词优先）。"""
    n = str(name).strip()
    n = _strip_leading_ascii_noise(n)
    # 品名常写「青岛全麦白啤」省略「啤酒」二字
    if n.startswith("青岛") and re.search(r"(全麦|白啤|奥古特|纯生|经典)", n):
        return "青岛啤酒"
    for p in KNOWN_BRAND_PREFIXES:
        if n.startswith(p):
            return p
    return ""


def _resolve_brand(raw: str, brand_field: str) -> str:
    """
    品牌：优先与商品名前缀一致的「商品品牌」字段；字段缺失、黑名单或与名称矛盾时，
    用名称前缀的已知品牌纠偏（如 名仁水蜜桃苏打水 纠正 ERP「圣诞」）。
    """
    b = re.sub(r"\[.*?\]", "", str(brand_field)).strip()
    if b in ("", "-", "未知", "nan"):
        b = ""
    inferred = _infer_brand_prefix(raw)
    if b and b in BRAND_FIELD_BLACKLIST:
        return inferred or b
    raw_norm = _strip_leading_ascii_noise(raw)
    # 子品牌归母品牌+子品牌，保证展示稳定
    if raw_norm.startswith("东方树叶"):
        return "农夫山泉东方树叶"
    if raw_norm.startswith("水溶C"):
        return "农夫山泉水溶C"
    if raw_norm.startswith("顺品郎"):
        return "郎酒"
    if b and raw_norm.startswith(b):
        return b
    if inferred:
        return inferred
    return b


def _strip_leading_wei_dairy_noise(work: str) -> str:
    """去掉乳品名里残留的孤字「味」（如切分错误产生的「味牛奶」）。"""
    if not work:
        return work
    return re.sub(
        r"^味+(?=(?:牛奶|酸奶|酸牛奶|发酵乳|风味发酵乳))",
        "",
        str(work).strip(),
    )


def _mother_brand(brand: str) -> str:
    b = re.sub(r"\[.*?\]", "", str(brand)).strip()
    return re.sub(r"-.*$", "", b).strip()


def _promote_subbrand(brand: str, work: str) -> tuple[str, str]:
    """
    母品牌字段匹配后，将子品牌从 work 头升格为展示品牌（伊利+畅轻→伊利畅轻；君乐宝+简醇→君乐宝简醇）。
    """
    if not brand or not work:
        return brand, work
    w = str(work).strip()
    mb = _mother_brand(brand)
    if mb == "伊利" and w.startswith("畅轻"):
        return "伊利畅轻", w[2:].lstrip()
    if mb == "君乐宝" and w.startswith("简醇"):
        return "君乐宝简醇", w[2:].lstrip()
    if mb == "新希望" and w.startswith("活润"):
        return "活润", w[2:].lstrip()
    return brand, w


def _dairy_changqing_fallback_body(brand: str, work: str, unified_cat: str) -> str:
    """畅轻杯装混合系列：无「发酵乳/酸奶」等后缀时，主体归为酸奶。"""
    if "畅轻" not in brand or unified_cat not in ("日配冷藏", "常温乳品"):
        return ""
    w = str(work).strip() if work else ""
    if not w:
        return ""
    if re.search(r"[+＋]|燕麦|爆珠|紫米|芦荟|青提|芭乐|石榴|莓|柚|红心|蓝莓|葡萄", w):
        return "酸奶"
    return ""


def refine_parsed_fields(
    raw_name: str,
    brand_field: str,
    uc: str,
    brand: str,
    body: str,
    flavor: str,
    package: str,
):
    """
    跨规则统一收口：包装长匹配、零度纠偏、方便面/乳品/冰淇淋/烤肠等易错 SKU。
    北道烧：多为调理肉制品（照烧/串烧类即食），「黑椒」等为口味而非主体的一部分。
    """
    raw = _normalize_product_title(str(raw_name).strip())
    raw_norm = _strip_leading_ascii_noise(raw)
    # Excel 合并后「统一大分类」常为 NaN → str 变成 "nan"，会导致所有按品类分支失效
    if uc is None or (isinstance(uc, float) and pd.isna(uc)):
        uc = ""
    else:
        uc = str(uc).strip()
    if uc.lower() in ("nan", "none"):
        uc = ""
    b = str(brand).strip()
    bd = str(body).strip()
    fl = str(flavor).strip()
    pk = str(package).strip()

    for p in sorted(PACKAGE_WORDS, key=len, reverse=True):
        if p in raw:
            pk = p
            break

    # 饮料：零度不是商品主体（覆盖表错误回填时亦在 main 中再做向量化兜底）
    if uc == "饮料" and bd and "零度" in bd:
        stripped = (
            bd.replace("零度", "")
            .replace("无糖", "")
            .replace("0糖", "")
            .replace("０糖", "")
            .strip()
        )
        if not stripped or stripped == "糖":
            bd = "可乐" if ("可乐" in raw_norm or "百事" in b or "pepsi" in raw_norm.lower()) else "碳酸饮料"
        else:
            bd = stripped

    if uc == "休闲零食":
        if "干吃奥尔良" in raw_norm or (bd == "干脆面" and "奥尔良" in raw_norm and not fl):
            bd = "干脆面"
            fl = fl or "奥尔良"
        if ("京门爆肚" in raw_norm or "爆肚" in raw_norm) and "麻辣" in raw_norm and not fl:
            fl = "麻辣"
        if "小公鱼" in raw_norm and "湘辣" in raw_norm and not fl:
            fl = "湘辣"
        if "麻肠" in raw_norm and "福字" in raw_norm:
            bd = "火腿肠"

    if uc == "方便速食":
        if "螺蛳粉" in raw_norm:
            if "好欢螺" in raw_norm:
                b = "好欢螺"
            bd = "螺蛳粉"
            if not fl and ("加辣" in raw_norm or "升级" in raw_norm):
                fl = "加辣"
        if "粉面菜蛋" not in raw_norm:
            if "大食桶" in raw_norm and (bd in ("方便面", "泡面", "汤面", "干拌面") or bd.startswith("方便面")):
                bd = "方便面大食桶"
            elif "大食袋" in raw_norm and (bd in ("方便面", "泡面") or bd.startswith("方便面")):
                bd = "方便面大食袋"
        if "粉面菜蛋" not in raw_norm:
            if "一袋半" in raw_norm:
                bd = "方便面"
                if "牛肉" in raw_norm and not fl:
                    fl = "牛肉"
            elif "一倍半" in raw_norm:
                bd = "方便面"
                if "牛肉" in raw_norm and not fl:
                    fl = "牛肉"
        if not fl or fl == "牛肉":
            noodle_map = (
                ("老坛酸菜牛肉", "酸菜牛肉"),
                ("酸菜牛肉", "酸菜牛肉"),
                ("老坛酸菜", "酸菜"),
                ("香辣牛肉", "香辣牛肉"),
                ("红烧牛肉", "红烧牛肉"),
                ("葱香排骨", "葱香排骨"),
                ("酸辣牛肉", "酸辣牛肉"),
                ("金汤肥牛", "金汤肥牛"),
            )
            for needle, lab in noodle_map:
                if needle in raw_norm:
                    fl = lab
                    break

    if uc in ("常温乳品", "日配冷藏"):
        dairy_skip = {"酸奶", "酸牛奶", "发酵乳", "AD钙奶", "优酸乳", "早餐奶", "奶茶", "奶酪棒", "奶酪益生元"}
        if bd not in dairy_skip and "早餐奶" not in raw_norm:
            if any(k in raw_norm for k in ("原生枕", "梦幻盖", "利乐枕")) and any(
                k in raw_norm for k in ("有机纯", "纯牛奶", "纯奶")
            ):
                bd = "纯牛奶"
            elif any(k in raw_norm for k in ("原生枕", "梦幻盖", "利乐枕", "利乐包")) and (
                "特仑苏" in raw_norm or "臻浓" in raw_norm or "臻享浓" in raw_norm
            ):
                bd = "牛奶"
        if "优酸乳" in raw_norm and raw_norm.rstrip().endswith("黄"):
            bd = "优酸乳"
            fl = fl or "黄桃"
        if "早餐奶" in raw_norm:
            bd = "早餐奶"
            if "核桃" in raw_norm:
                fl = fl or "核桃"
        if bd == "液态奶" and ("臻浓" in raw_norm or "臻享浓" in raw_norm):
            bd = "牛奶"
        # 品牌词误入主体（如 D科迪原生枕180g）
        if bd in ("科迪", "科迪-低温"):
            bd = "纯牛奶"

    if uc == "冷冻食品" and ("巧乐兹" in raw_norm or b == "巧乐兹" or "巧乐兹" in bd):
        if "脆筒" in raw_norm or "脆筒" in bd:
            bd = "冰淇淋脆筒"
        elif "巧脆棒" in raw_norm or "脆棒" in raw_norm or "巧脆棒" in bd:
            bd = "冰淇淋巧脆棒"
        if not fl:
            if "香草" in raw_norm:
                fl = "香草味"
            elif "巧克力" in raw_norm:
                fl = "巧克力味"
            elif "草莓" in raw_norm:
                fl = "草莓味"

    if uc == "FF速食":
        # 覆盖表常把「黑椒鸡排」整段写进主体，需拆开
        m_ff = re.match(r"^(黑椒|奥尔良|香辣|蒜香|孜然|蜜汁)(鸡排|鸡腿|鸡块|鸡翅)$", bd)
        if m_ff:
            bd = m_ff.group(2)
            fl = fl or m_ff.group(1)
        if bd == "鸡排" or bd.endswith("鸡排"):
            if "黑椒" in raw_norm and not fl:
                fl = "黑椒"
            elif "奥尔良" in raw_norm and not fl:
                fl = "奥尔良"
        if "地道大肉肠" in raw_norm or bd == "大肉肠":
            bd = "烤肠"

    if bd == "北道烧" and "黑椒" in raw_norm and not fl:
        fl = "黑椒"

    # ── 品名强收口（不依赖统一大分类：类目映射错误/NaN 时仍纠正）────────
    _rn = raw_norm
    if "巧乐兹" in _rn:
        if "脆筒" in _rn or "脆筒" in bd:
            bd = "冰淇淋脆筒"
        elif "巧脆棒" in _rn or "脆棒" in _rn or "巧脆棒" in bd:
            bd = "冰淇淋巧脆棒"
        if not fl:
            if "香草" in _rn:
                fl = "香草味"
            elif "巧克力" in _rn:
                fl = "巧克力味"
            elif "草莓" in _rn:
                fl = "草莓味"
        if (not b or b.lower() == "nan") and _rn.startswith("巧乐兹"):
            b = "巧乐兹"
    if "螺蛳粉" in _rn:
        if "好欢螺" in _rn:
            b = "好欢螺"
        bd = "螺蛳粉"
        if not fl and ("加辣" in _rn or "升级" in _rn):
            fl = "加辣"
    if "螺蛳粉" not in _rn and "粉面菜蛋" not in _rn:
        if "大食桶" in _rn and (not bd.startswith("方便面") or bd in ("方便面", "泡面")):
            bd = "方便面大食桶"
        elif "大食袋" in _rn and (not bd.startswith("方便面") or bd in ("方便面", "泡面")):
            bd = "方便面大食袋"
        if "一袋半" in _rn:
            bd = "方便面"
            if "牛肉" in _rn and not fl:
                fl = "牛肉"
        elif "一倍半" in _rn:
            bd = "方便面"
            if "牛肉" in _rn and not fl:
                fl = "牛肉"
    if "螺蛳粉" not in _rn and (not fl or fl == "牛肉"):
        for needle, lab in (
            ("老坛酸菜牛肉", "酸菜牛肉"),
            ("酸菜牛肉", "酸菜牛肉"),
            ("老坛酸菜", "酸菜"),
            ("香辣牛肉", "香辣牛肉"),
            ("红烧牛肉", "红烧牛肉"),
            ("葱香排骨", "葱香排骨"),
            ("酸辣牛肉", "酸辣牛肉"),
            ("金汤肥牛", "金汤肥牛"),
        ):
            if needle in _rn:
                fl = lab
                break
    # 乳品：品名含包型+纯奶语义即收（不论大类是否写成常温乳品）
    if "早餐奶" in _rn:
        bd = "早餐奶"
        if "核桃" in _rn:
            fl = fl or "核桃"
    elif "优酸乳" in _rn and _rn.rstrip().endswith("黄"):
        bd = "优酸乳"
        fl = fl or "黄桃"
    elif (
        any(k in _rn for k in ("梦幻盖", "原生枕", "利乐枕", "利乐包"))
        and "酸奶" not in bd
        and "酸牛奶" not in _rn
        and "优酸乳" not in bd
    ):
        if any(k in _rn for k in ("有机纯", "纯牛奶", "纯奶", "金典有机", "金典")):
            bd = "纯牛奶"
        elif any(k in _rn for k in ("特仑苏", "臻浓", "臻享浓")) and "牛奶" in _rn:
            bd = "牛奶"
    # 鸡排：品名同时含口味词与鸡排
    if "三明治" not in _rn and "鸡排" in _rn:
        m2 = re.match(r"^(黑椒|奥尔良|香辣|蒜香|孜然|蜜汁)(鸡排|鸡腿|鸡块|鸡翅)$", bd)
        if m2:
            bd, fl = m2.group(2), fl or m2.group(1)
        elif "黑椒" in _rn and bd != "鸡排":
            bd, fl = "鸡排", fl or "黑椒"
        elif "奥尔良" in _rn and bd != "鸡排":
            bd, fl = "鸡排", fl or "奥尔良"
    if ("地道大肉肠" in _rn or bd == "大肉肠") and "三明治" not in _rn:
        bd = "烤肠"

    # ── 复合主体拆写（品名/覆盖表整段当主体时）────────────────────────
    if bd in ("臻享浓牛奶", "臻浓牛奶") or re.match(r"^臻享?浓牛奶$", bd):
        bd = "牛奶"
    if "果粒优酸乳" in _rn or "果粒优酸乳" in bd:
        bd = "优酸乳"
        if "蓝莓" in _rn:
            fl = fl or "蓝莓"
        elif _rn.rstrip().endswith("黄") or "优酸乳黄" in _rn or bd.endswith("黄"):
            fl = fl or "黄桃"
    elif "优酸乳" in _rn and "蓝莓" in _rn and "果粒" not in _rn:
        bd = "优酸乳"
        fl = fl or "蓝莓"
    if "未来星" in _rn:
        b = b or "未来星"
        if "儿童" in _rn and "有机" in _rn:
            bd = "儿童有机奶"
        elif "儿童" in _rn:
            bd = "儿童牛奶"
        else:
            bd = "牛奶"
    if "礼拜天" in _rn:
        b = b or "礼拜天"
        if "方糕" in _rn:
            bd = "方糕"
            if "红豆" in _rn:
                fl = fl or "红豆"
    if ("牛乳" in _rn or "牛乳" in bd) and ("脆筒" in _rn or "脆筒" in bd):
        bd = "冰淇淋脆筒"
        fl = fl or "牛乳味"
    if bd == "牛乳口味脆筒" or ("口味脆筒" in bd and "牛乳" in bd):
        bd = "冰淇淋脆筒"
        fl = fl or "牛乳味"
    if "北道烧" in _rn or "北道烧" in bd:
        if bd != "北道烧" and "北道烧" in bd:
            if "黑椒" in bd or "黑椒" in _rn:
                fl = fl or "黑椒"
            bd = "北道烧"
    if "京门爆肚麻辣" in bd or ("京门爆肚" in _rn and "麻辣" in _rn):
        bd, fl = "素毛肚", fl or "麻辣"
    if "湘辣小公鱼" in bd or ("小公鱼" in _rn and "湘辣" in _rn):
        bd, fl = "小鱼干", fl or "湘辣"
    if ("咪咪" in _rn or "咪咪" in bd) and ("虾条" in _rn or "虾条" in bd):
        b = b or "咪咪"
        bd = "虾条"
        if "虾味" in _rn or "虾味" in bd:
            fl = fl or "虾味"
    if "乐事" in _rn or "乐事" in bd:
        if "黄瓜" in _rn or "黄瓜" in bd or "黄瓜味" in bd:
            b = b or "乐事"
            bd = "薯片"
            fl = fl or "黄瓜味"
    m_gan = re.match(r"^干吃(.+)$", bd)
    if m_gan and "干吃面" not in bd:
        tail = m_gan.group(1)
        bd = "干脆面"
        if tail.endswith("味") and not fl:
            fl = tail
        elif "鸡翅" in tail and not fl:
            fl = "鸡翅味"
    if bd == "麻肠" or (bd.startswith("麻肠") and "福字" not in _rn):
        bd = "火腿肠"

    bd = _strip_body_packaging_tokens(bd)
    bd = _strip_brand_from_body(bd, b)
    # 常见「主体带口味」拆分：如 京门爆肚麻辣 / 乐事黄瓜味 / 咪咪虾条虾味
    m_tail_flv = re.match(r"^(.+?)(黑椒|奥尔良|香辣|麻辣|黄瓜味|虾味|蓝莓味|牛乳味|牛奶味|红豆)$", bd)
    if m_tail_flv:
        head, tail_flv = m_tail_flv.group(1), m_tail_flv.group(2)
        if head and len(head) >= 2:
            bd = head
            fl = fl or tail_flv
    # 方便面形态名统一到主体/包装分离
    if bd in ("方便面一袋半", "方便面一倍半"):
        bd = "方便面"
    # 杯面统一入方便面主体
    if "杯面" in bd:
        bd = "方便面"
    # 针对你反馈的典型复合主体做非品牌化收口
    if bd in ("邬辣妈素牛筋", "素牛筋"):
        bd = "素牛筋"
    if "脆香米" in bd:
        bd = "巧克力"
    if any(k in bd for k in ("冰球杯", "袋冰", "食用袋冰")):
        bd = "冰杯"
    if "老中街冰棍" in bd or "冰棍" in bd:
        bd = "冰棍"
    if "俄式大蛋筒" in bd or "蛋筒" in bd:
        bd = "冰淇淋蛋筒"
    if bd == "康美真果粒" or "真果粒" in bd:
        bd = "果粒乳饮料"
    if bd in ("金典有机奶", "有机奶"):
        bd = "纯牛奶"
    if "全脂奶" in bd:
        bd = "全脂牛奶"
    if "日式豚骨杯面" in bd:
        bd = "方便面"
        fl = fl or "日式豚骨"
    # 自热/自煮火锅统一（如「小酥肉自煮火锅」）
    if "自煮火锅" in bd or "自热火锅" in bd or "自煮火锅" in _rn or "自热火锅" in _rn:
        bd = "自热火锅"
        if ("小酥肉" in _rn or "小酥肉" in str(raw_name)) and not fl:
            fl = "小酥肉"
    # 你反馈的主体污染样式：品牌+主体/乱码切分/型号残片
    if bd in ("安慕希高端凝酪", "高端凝酪"):
        bd = "凝酪酸奶"
    if bd in ("谷粒多燕麦牛奶",):
        bd = "燕麦牛奶"
    if bd in ("锦甜墨鱼爆蛋",) or "锦甜墨鱼爆蛋" in _rn or ("墨鱼爆蛋" in _rn and "锦甜" in _rn):
        b = b or "锦甜"
        bd = "墨鱼爆蛋"
    if "酸奶棒糖" in _rn:
        bd = "酸奶棒糖"
    elif "酸奶奶棒" in _rn:
        bd = "酸奶奶棒"
    if bd in ("喜力",):
        # 仅主体异常兜底；酒类品牌词落主体时回归啤酒
        bd = "啤酒"
    if bd in ("汁1", "汁", "饮1"):
        # 根因修复：如「红苹果汁1L」在抽走“红苹果”后残留为“汁1”
        if "红苹果汁" in _rn:
            bd = "红苹果汁"
            if fl == "红苹果":
                fl = "苹果"
        elif "苹果汁" in _rn:
            bd = "苹果汁"
            if fl in ("红苹果", "青苹果"):
                fl = "苹果"
        elif "果汁" in _rn:
            bd = "果汁饮料"
        elif "汁" in _rn:
            bd = "果蔬汁"
        else:
            bd = "饮料"
    if b.lower() in ("nan", "none"):
        b = ""
    return b, bd, fl, pk


def apply_parse_refinements(
    df: pd.DataFrame,
    name_col: str = "商品名称",
    brand_field_col: str = "商品品牌",
    uc_col: str = "统一大分类",
    out_brand_col: str = "品牌_解析",
    out_body_col: str = "商品主体",
    out_flavor_col: str = "口味",
    out_pkg_col: str = "包装类型",
) -> None:
    """
    对整表四列（品牌/主体/口味/包装）执行 refine_parsed_fields。
    必须在「覆盖表回填 + 规则解析」之后调用，否则覆盖表里的历史错行会一直带到页面/JSON。
    """
    need = [name_col, uc_col, out_brand_col, out_body_col, out_flavor_col, out_pkg_col]
    for c in need:
        if c not in df.columns:
            return
    if brand_field_col not in df.columns:
        df[brand_field_col] = ""

    def _cell(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        s = str(v).strip()
        return "" if s.lower() == "nan" else s

    def _one_row(r):
        return refine_parsed_fields(
            _cell(r[name_col]) or str(r[name_col]),
            _cell(r.get(brand_field_col)),
            _cell(r.get(uc_col)),
            _cell(r.get(out_brand_col)),
            _cell(r.get(out_body_col)),
            _cell(r.get(out_flavor_col)),
            _cell(r.get(out_pkg_col)),
        )

    res = df.apply(
        lambda r: pd.Series(
            _one_row(r),
            index=[out_brand_col, out_body_col, out_flavor_col, out_pkg_col],
        ),
        axis=1,
    )
    df.loc[:, [out_brand_col, out_body_col, out_flavor_col, out_pkg_col]] = res


def parse_product(name, brand_field="", unified_cat=""):
    """
    解析商品名称 → 品牌_解析 / 商品主体 / 口味 / 规格数值 / 规格单位 / 包装类型

    解析逻辑：
      品牌：字段与名称前缀一致则用字段；否则用已知品牌前缀（含「元气森林外星人」）
      度数（酒类）→ 归入口味字段（不是规格）
      商品主体：先在 work 上按品类锁定 CAT_PRODUCT_TYPES / PRODUCT_TYPES（最长优先由词典顺序保证）
      口味：遮蔽已锁定的主体子串后再按 FLAVOR_WORDS 匹配；仍无则可用 BODY_TO_FLAVOR_HINT[主体] 补全
      未锁定主体时：沿用「去口味余量→再匹配类型」的兜底逻辑
    """
    raw = _normalize_product_title(str(name).strip())
    raw_norm = _strip_leading_ascii_noise(raw)
    brand = _resolve_brand(raw, brand_field)

    # 1. 提取度数
    degree = ""
    dm = DEGREE_RE.search(raw_norm)
    if dm:
        degree = dm.group(1) + "°"

    # 2. 去度数 → 去品牌前缀，得到待拆分的中间串
    work = DEGREE_RE.sub("", raw_norm).strip()
    if brand:
        work = re.sub(r"^" + re.escape(brand) + r"\s*", "", work).strip()
    brand, work = _promote_subbrand(brand, work)

    # 去规格后再做主体/口味匹配（避免 100g 打断「起酥苹果」等）
    work_match = SPEC_RE.sub("", work).strip() if work else ""
    work_match = _strip_leading_wei_dairy_noise(work_match)

    body_label, body_span = "", ""
    flavor = ""
    if unified_cat is None or (isinstance(unified_cat, float) and pd.isna(unified_cat)):
        uc = ""
    else:
        uc = str(unified_cat).strip()
    if uc.lower() in ("nan", "none"):
        uc = ""

    bl_m, bs_m, fl_m = _try_mooncake_sugar_prefix(work_match)
    if bs_m:
        body_label, body_span, flavor = bl_m, bs_m, fl_m
    else:
        body_label, body_span = _lock_body_suffix_first(work_match, unified_cat)
        if not body_label:
            body_label, body_span = _lock_body_label(work_match, unified_cat)
        if not body_label and brand and "锐澳" in brand and uc == "酒类":
            body_label = "鸡尾酒"
        if not body_label:
            fb = _dairy_changqing_fallback_body(brand, work_match, uc)
            if fb:
                body_label = fb

    # 纠错：如果“主体”长得像纯口味词（xx味），且当前尚未识别到口味，
    # 则很可能是主体/口味切分反了（例如：饮料把“香草味”当成主体）。
    if body_label and not flavor and re.fullmatch(r".+味", str(body_label).strip()):
        flavor = str(body_label).strip()
        body_label = ""
        body_span = ""

    work_for_flavor = work_match
    if body_span:
        work_for_flavor = work_for_flavor.replace(body_span, "\uff03" * len(body_span), 1)

    if not flavor:
        for f in FLAVOR_WORDS_SORTED:
            if f in work_for_flavor:
                flavor = f
                break
    if not flavor and degree:
        flavor = degree
    if not flavor and body_label:
        hint = BODY_TO_FLAVOR_HINT.get(body_label)
        if hint:
            flavor = hint
    if not flavor and body_label in ("红苹果汁", "苹果汁") and body_span and "苹果" in body_span:
        flavor = "苹果"
    if not flavor and body_label == "起酥面包" and body_span and "苹果" in body_span:
        flavor = "苹果"

    # 4. 规格（容量/重量）
    spec_val = spec_unit = ""
    m = SPEC_RE.search(raw)
    if m:
        spec_val, spec_unit = m.group(1), m.group(2)

    # 5. 包装类型（长词优先，避免只命中「袋」等短词）
    package = ""
    for p in sorted(PACKAGE_WORDS, key=len, reverse=True):
        if p in raw:
            package = p
            break

    # 6. 商品主体
    if body_label:
        body = body_label
    else:
        work_for_type = work
        if flavor and flavor != degree:
            work_for_type = work.replace(flavor, "", 1).strip()

        product_type = ""
        cat_patterns = CAT_PRODUCT_TYPES.get(str(unified_cat).strip(), [])
        for pattern, label in cat_patterns:
            if re.search(pattern, work_for_type):
                product_type = label
                break
        if not product_type:
            for pattern, label in PRODUCT_TYPES:
                if re.search(pattern, work_for_type):
                    product_type = label
                    break

        if product_type:
            body = product_type
        else:
            body = work_for_type
            body = SPEC_RE.sub("", body)
            if flavor and flavor != degree:
                body = body.replace(flavor, "")
            for p in PACKAGE_WORDS:
                body = body.replace(p, "")
            body = re.sub(r"[A-Za-z0-9]{3,}$", "", body)
            body = re.sub(r"[（(][^）)]*[）)]", "", body)
            body = re.sub(r"[\s\-\_\.]+", "", body).strip()
            if body in BODY_REMAINDER_MAP:
                body = BODY_REMAINDER_MAP[body]
            if len(body) <= 1:
                remn = work_for_type.strip()
                if remn in PRODUCT_LINE_ONLY_MARKERS or remn in PRODUCT_LINE_ONLY_MARKERS_2:
                    body = _coerce_body_if_only_product_line(remn, brand, raw_norm, unified_cat)
                elif remn and len(remn) <= 4:
                    # 避免主体退化成纯规格（如“38g”“45g”）
                    if re.match(r"^\d+(?:\.\d+)?\s*[A-Za-z一-龥]+$", remn):
                        if uc == "休闲零食":
                            body = "休闲食品"
                        elif uc == "日化美护":
                            body = "日化用品"
                        else:
                            body = brand if brand else raw
                    else:
                        body = remn
                elif re.match(r"^\d+(?:\.\d+)?\s*[A-Za-z一-龥]+$", remn):
                    if uc == "休闲零食":
                        body = "休闲食品"
                    elif uc == "日化美护":
                        body = "日化用品"
                    else:
                        body = brand if brand else raw
                elif remn and len(remn) <= 6 and uc == "休闲零食":
                    body = "休闲食品"
                else:
                    body = brand if brand else raw

    body = _coerce_body_if_only_product_line(body, brand, raw_norm, unified_cat)

    # 利乐枕/梦幻盖等包装词不应残留在主体上
    body = _strip_body_packaging_tokens(body)

    # 新希望「荷荷」系列：荷荷为产品线名，主体应为含乳饮料
    if "荷荷" in brand and uc in ("日配冷藏", "常温乳品"):
        body = "含乳饮料"
        if not flavor:
            for w in ("荔枝", "草莓", "葡萄", "白桃", "柚子"):
                if w in work_match:
                    flavor = w
                    break

    # 安全套品类常见写法不带“避孕套”字样（如 杜蕾斯love大胆爱吧）
    if uc == "日化美护" and any(k in str(brand) for k in ("杜蕾斯", "冈本")):
        body = "避孕套"

    # 休闲零食里部分品牌名会被先剥离，导致主体仅剩口味词，做品牌级兜底
    if uc == "休闲零食" and "都市牧场" in str(brand) and body in ("休闲食品", "含片糖", ""):
        body = "含片"
    if uc == "休闲零食" and "好友趣" in str(brand) and body in ("休闲食品", "膨化食品", ""):
        body = "薯片"

    # refine 在 main 中通过 apply_parse_refinements 对「覆盖表+规则」全表统一执行，避免命中覆盖时跳过纠偏
    return pd.Series([brand, body, flavor, spec_val, spec_unit, package])


# ══════════════════════════════════════════════════
# 热度计算
# ══════════════════════════════════════════════════

def calc_tag_heat(df, tag_col, dim_label=None):
    """
    计算标签维度热度

    热度评分 = 当前量级×40% + 时间趋势×35% + 地域广度×25%
      当前量级：销量PSD(25%) + 铺货转化率(10%) + 客数PSD(5%)
      时间趋势：同标签跨期销量PSD线性回归斜率（单期→0.5，不奖不罚）
      地域广度：该标签覆盖的地区数

    评分在每个(品类×标签维度)内归一化，仅用于同类横向比较。

    dim_label：写入结果中的「标签维度」展示名；默认与 tag_col 相同。
              例如 tag_col 用「口味_统计」归一列聚合，dim_label 仍填「口味」供前端筛选。
    """
    out_dim = dim_label if dim_label is not None else tag_col
    valid = df[
        df[tag_col].notna() &
        (df[tag_col].astype(str).str.strip() != "") &
        (df[tag_col].astype(str) != "nan")
    ].copy()
    if valid.empty:
        return pd.DataFrame()

    rows = []
    for (cat, tag), g in valid.groupby(["统一大分类", tag_col]):

        # 时间趋势（按库存店数加权的期间PSD均值，跨期线性回归）
        g2 = g.copy()
        g2["_w"] = np.where(g2["库存店数"] > 0, g2["库存店数"], 1)
        g2["_wpsd"] = g2["销量PSD"] * g2["_w"]
        by_period = (
            g2.groupby("数据期间", as_index=False)[["_wpsd", "_w"]].sum()
            .assign(期PSD=lambda x: x["_wpsd"] / x["_w"])
            .sort_values("数据期间")
        )
        n_p = len(by_period)
        if n_p >= 2:
            x = np.arange(n_p, dtype=float)
            y = by_period["期PSD"].values.astype(float)
            mean_y = y.mean() if y.mean() != 0 else 1.0
            slope = np.polyfit(x, y, 1)[0] / mean_y
            trend_score = float(np.clip((slope + 1) / 2, 0, 1))
            trend_label = ("↑ 上升" if slope > 0.1 else
                           "↓ 下滑" if slope < -0.1 else "→ 平稳")
        else:
            trend_score = 0.5
            trend_label = "— 单期"

        rep_brand = ""
        if out_dim == "商品主体" and "品牌_解析" in g.columns and "销量" in g.columns:
            gb = g.copy()
            gb["品牌_解析"] = gb["品牌_解析"].astype(str).str.strip()
            gb = gb[gb["品牌_解析"].notna() & (gb["品牌_解析"] != "") & (gb["品牌_解析"] != "nan")]
            if not gb.empty:
                rep_brand = str(gb.groupby("品牌_解析")["销量"].sum().idxmax())

        rows.append({
            "统一大分类":     cat,
            "标签维度":       out_dim,
            "标签值":         str(tag),
            "代表品牌":       rep_brand,
            "SKU数量":        g["商品名称"].nunique(),
            "覆盖地区数":     g["地域名称"].nunique(),
            "平均销量PSD":    round(g["销量PSD"].mean(), 3),
            "平均铺货转化率": round(g["铺货转化率"].mean(), 3),
            "平均毛利率":     round(g["毛利率"].mean(), 3),
            "平均客数PSD":    round(g["客数PSD"].mean(), 3),
            "总销量":         int(g["销量"].sum()),
            "趋势方向":       trend_label,
            "_t":  trend_score,
            "_p":  g["销量PSD"].mean(),
            "_c":  g["铺货转化率"].mean(),
            "_cl": g["客数PSD"].mean(),
            "_r":  float(g["地域名称"].nunique()),
        })

    result = pd.DataFrame(rows)

    # (品类×维度)内归一化，用 transform 避免 groupby 丢列
    for src, dst in [("_p","_np"), ("_c","_nc"), ("_cl","_ncl"), ("_t","_nt"), ("_r","_nr")]:
        result[dst] = result.groupby(["统一大分类", "标签维度"])[src].transform(
            lambda s: ((s - s.min()) / (s.max() - s.min())).clip(0, 1)
            if s.max() > s.min() else 0.5
        )

    result["热度评分"] = (
        (result["_np"] * 0.25 + result["_nc"] * 0.10 + result["_ncl"] * 0.05) * 0.40
        + result["_nt"] * 0.35
        + result["_nr"] * 0.25
    ).round(4)

    result["热度等级"] = result["热度评分"].apply(
        lambda s: "🔥 高热" if s >= 0.75 else
                  "📈 上升" if s >= 0.50 else
                  "➡️ 平稳" if s >= 0.25 else "❄️ 冷淡"
    )

    # 机会信号：销量PSD高但SKU少 = 供给不足，开发机会
    psd_q60 = result["平均销量PSD"].quantile(0.6)
    result["机会信号"] = result.apply(
        lambda r: "⚡ 高需低供" if (r["平均销量PSD"] > psd_q60 and r["SKU数量"] <= 2)
                  else "", axis=1
    )

    return result


# ══════════════════════════════════════════════════
# Excel 样式
# ══════════════════════════════════════════════════

def make_fill(hex6):
    return PatternFill("solid", start_color=hex6)

FILLS = {
    "header": make_fill("1F4E79"),
    "new":    make_fill("E2EFDA"),
    "hot":    make_fill("FCE4D6"),
    "rise":   make_fill("E2EFDA"),
    "flat":   make_fill("EDEDED"),
    "fall":   make_fill("FFE0E0"),
    "stripe": make_fill("EBF3FB"),
    "score":  make_fill("FFF2CC"),
}
H_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
N_FONT = Font(name="Arial", size=10)
B_FONT = Font(name="Arial", size=10, bold=True)
C_ALIGN = Alignment(horizontal="center", vertical="center")
L_ALIGN = Alignment(horizontal="left",   vertical="center")
THIN = Border(**{s: Side(style="thin", color="CCCCCC")
                 for s in ("left", "right", "top", "bottom")})

def hdr(cell, key="header"):
    cell.font = H_FONT; cell.fill = FILLS[key]
    cell.alignment = C_ALIGN; cell.border = THIN

def cel(cell, key=None, bold=False):
    cell.font = B_FONT if bold else N_FONT
    cell.border = THIN; cell.alignment = L_ALIGN
    if key: cell.fill = FILLS[key]

def write_heat_sheet(ws, data, out_cols):
    col_w = {
        "统一大分类": 12, "标签维度": 10, "标签值": 16, "代表品牌": 12,
        "热度评分": 10, "热度等级": 10, "趋势方向": 10, "机会信号": 12,
        "SKU数量": 8, "覆盖地区数": 10, "平均销量PSD": 12,
        "平均铺货转化率": 12, "平均毛利率": 10, "平均客数PSD": 10, "总销量": 10,
    }
    for ci, col in enumerate(out_cols, 1):
        hdr(ws.cell(row=1, column=ci, value=col))
    prev_cat = prev_dim = None
    score_idx = out_cols.index("热度评分")
    for ri, row in enumerate(dataframe_to_rows(data, index=False, header=False), 2):
        cur_cat, cur_dim = row[0], row[1]
        score = row[score_idx]
        for ci, val in enumerate(row, 1):
            cn = out_cols[ci - 1]
            if cn == "热度评分":
                k = ("hot"  if score >= 0.75 else
                     "rise" if score >= 0.50 else
                     "flat" if score >= 0.25 else "fall")
            elif cur_cat != prev_cat or cur_dim != prev_dim:
                k = "stripe"
            else:
                k = None
            cel(ws.cell(row=ri, column=ci, value=val), k, bold=(cn == "热度评分"))
        prev_cat = cur_cat; prev_dim = cur_dim
    for ci, col in enumerate(out_cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_w.get(col, 12)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ══════════════════════════════════════════════════
# 主流程
# ══════════════════════════════════════════════════

def main():
    print("=" * 55)
    print("  快消品标签热度分析脚本")
    print("=" * 55)

    # ── 1. 读取 ────────────────────────────────────
    print("\n[1/6] 读取数据...")
    dfs = []
    for source, path in INPUT_FILES.items():
        df_tmp = read_file(_abs_project_path(path), source, sep=FILE_SEP)
        if source == "渠道A":
            df_tmp = df_tmp.rename(columns=CHANNEL_A_RENAME)
        dfs.append(df_tmp)
        print(f"      {source}: {len(df_tmp):,} 行  ←  {path}")
    df = pd.concat(dfs, ignore_index=True)
    print(f"      合并后共 {len(df):,} 行")

    # ── 2. 分类映射 ────────────────────────────────
    print("\n[2/6] 统一大分类映射...")
    df["大分类名称"] = df["大分类名称"].astype(str).str.strip()
    df["统一大分类"] = df.apply(
        lambda r: CAT_MAP.get((r["渠道来源"], r["大分类名称"]), r["大分类名称"]),
        axis=1
    )

    _n_cat = len(df)
    df = df[df["统一大分类"] != "现制品"].copy()
    if len(df) < _n_cat:
        print(f"      已剔除「现制品」{_n_cat - len(df):,} 行")

    # ── 3. 地域解析 ────────────────────────────────
    print("[3/6] 地域解析...")
    df[["省份", "城市", "渠道类型"]] = df["地域名称"].apply(parse_region)

    # ── 4. 商品名称解析 ────────────────────────────
    print("[4/6] 商品名称解析（数据量大时耗时较长）...")
    # 解析结果优先使用覆盖表（半自动/人工补全沉淀），未命中才走 parse_product 规则解析
    override_path = _abs_project_path("商品解析覆盖表.xlsx")

    # 用于 join_key_no_code 的名称归一化（与 build_product_master.py 保持一致）
    # 注意：这里在后面使用向量化实现，函数主要保留给可读性
    def _normalize_name_for_join(s: str) -> str:
        t = str(s).strip()
        t = re.sub(r"^[【\[]\s*新\s*[】\]]\s*", "", t)
        t = re.sub(r"^[（(]\s*新\s*[）)]\s*", "", t)
        t = re.sub(r"^[^一-龥A-Za-z0-9]+", "", t)
        t = re.sub(r"^[A-Za-z]{1,3}\s*", "", t)
        return t.strip()

    ov = None
    ov_map = {}
    if os.path.exists(override_path):
        try:
            ov = pd.read_excel(override_path, sheet_name="master_overrides")
        except Exception:
            ov = None

    if ov is not None and not ov.empty:
        key_col = "join_key_no_code" if "join_key_no_code" in ov.columns else "sku_key"

        def _clean_str(v):
            if pd.isna(v):
                return ""
            return str(v).strip()

        def _clean_spec_val(v):
            if pd.isna(v):
                return ""
            # Excel 可能把整数规格写成 100.0
            if isinstance(v, (int,)):
                return str(v)
            if isinstance(v, float):
                if v.is_integer():
                    return str(int(v))
                return str(v)
            sv = str(v).strip()
            return sv

        ov = ov.copy()
        ov["__key"] = ov[key_col].apply(_clean_str)
        for fld in ["brand", "body", "flavor", "spec_val", "spec_unit", "package", "confidence"]:
            if fld not in ov.columns:
                ov[fld] = ""

        ov_map = {
            row["__key"]: {
                "brand": _clean_str(row["brand"]),
                "body": _clean_str(row["body"]),
                "flavor": _clean_str(row["flavor"]),
                "spec_val": _clean_spec_val(row["spec_val"]),
                "spec_unit": _clean_str(row["spec_unit"]),
                "package": _clean_str(row["package"]),
            }
            for _, row in ov.iterrows()
            if row.get("__key", "").strip() != ""
        }

    # join_key_no_code：渠道 + 归一化商品名称（没有商品编码时的替代主键）
    _name = df["商品名称"].astype(str).str.strip()
    _name = _name.str.replace(r"^[【\[]\s*新\s*[】\]]\s*", "", regex=True)
    _name = _name.str.replace(r"^[（(]\s*新\s*[）)]\s*", "", regex=True)
    _name = _name.str.replace(r"^[^一-龥A-Za-z0-9]+", "", regex=True)
    _name = _name.str.replace(r"^[A-Za-z]{1,3}\s*", "", regex=True)
    _name = _name.str.strip()
    _ch = df["渠道来源"].astype(str).str.strip().replace({"渠道A": "壹度", "渠道B": "安达"})
    df["__join_key_no_code"] = _ch + "::N::" + _name
    hit_mask = df["__join_key_no_code"].isin(ov_map.keys()) if ov_map else pd.Series([False] * len(df), index=df.index)

    parse_cols = ["品牌_解析", "商品主体", "口味", "规格数值", "规格单位", "包装类型"]
    # 默认先留空
    for c in parse_cols:
        df[c] = ""

    if hit_mask.any():
        # 命中覆盖表：直接回填（避免逐行 apply，使用 dict+map）
        brand_dict = {k: v.get("brand", "") for k, v in ov_map.items()}
        body_dict = {k: v.get("body", "") for k, v in ov_map.items()}
        flavor_dict = {k: v.get("flavor", "") for k, v in ov_map.items()}
        spec_val_dict = {k: v.get("spec_val", "") for k, v in ov_map.items()}
        spec_unit_dict = {k: v.get("spec_unit", "") for k, v in ov_map.items()}
        package_dict = {k: v.get("package", "") for k, v in ov_map.items()}

        k = df.loc[hit_mask, "__join_key_no_code"]
        df.loc[hit_mask, "品牌_解析"] = k.map(brand_dict).fillna("")
        df.loc[hit_mask, "商品主体"] = k.map(body_dict).fillna("")
        df.loc[hit_mask, "口味"] = k.map(flavor_dict).fillna("")
        df.loc[hit_mask, "规格数值"] = k.map(spec_val_dict).fillna("")
        df.loc[hit_mask, "规格单位"] = k.map(spec_unit_dict).fillna("")
        df.loc[hit_mask, "包装类型"] = k.map(package_dict).fillna("")

    # 未命中覆盖表：走原规则解析
    miss_mask = ~hit_mask
    if miss_mask.any():
        parsed = df.loc[miss_mask].apply(
            lambda r: parse_product(r["商品名称"], r.get("商品品牌", ""), r.get("统一大分类", "")),
            axis=1,
        )
        parsed.columns = parse_cols
        # 用 values 进行位置赋值，避免 parsed 的 index 与 df.loc 子集不一致导致覆盖命中行
        df.loc[miss_mask, parse_cols] = parsed.values

    # 覆盖表命中行不会走 parse_product，必须在此对全表统一 refine，页面/JSON 才与规则一致
    print("      全表解析收口 apply_parse_refinements（含覆盖表命中行）...")
    apply_parse_refinements(df)

    # 全表兜底：饮料类「零度」不得作为商品主体（含覆盖表历史错误）
    _ling = df["统一大分类"].astype(str).eq("饮料") & df["商品主体"].astype(str).str.contains(
        "零度", na=False
    )
    if _ling.any():
        _nm = df.loc[_ling, "商品名称"].astype(str)
        df.loc[_ling, "商品主体"] = np.where(_nm.str.contains("可乐", na=False), "可乐", "碳酸饮料")

    # 清理临时列
    if "__join_key_no_code" in df.columns:
        df.drop(columns=["__join_key_no_code"], inplace=True)

    # ── 5. 数值清洗 ────────────────────────────────
    print("[5/6] 数值清洗与指标计算...")
    for col in ["销量PSD", "销售额PSD", "客数PSD",
                "动销店数", "库存店数", "毛利率",
                "销量", "销售额", "客数", "售价"]:
        if col in df.columns:
            df[col] = pd.to_numeric(
                df[col].astype(str).str.replace("%", "").str.strip(),
                errors="coerce"
            ).fillna(0)

    # 毛利率可能混合两种口径：0-1 与 0-100，按行归一化
    df["毛利率"] = df["毛利率"].apply(lambda x: x / 100 if x > 1 else x)

    df["铺货转化率"] = df.apply(
        lambda r: round(r["动销店数"] / r["库存店数"], 4) if r["库存店数"] > 0 else 0,
        axis=1
    )

    df["价格带"] = pd.cut(
        df["售价"], bins=PRICE_BINS, labels=PRICE_LABELS, right=True
    ).astype(str).replace("nan", "未知价格")

    # 统计口径口味（合并「苹果味/苹果/红苹果」等）；热度与 JSON 用此列，宽表仍保留原「口味」便于下钻
    if "口味" in df.columns:
        df["口味_统计"] = df["口味"].apply(normalize_flavor_for_stats)
    else:
        df["口味_统计"] = pd.Series([""] * len(df), index=df.index, dtype=object)

    if "商品主体" in df.columns:
        df["商品主体_统计"] = df["商品主体"].apply(normalize_body_for_stats)
    else:
        df["商品主体_统计"] = pd.Series([""] * len(df), index=df.index, dtype=object)

    # ── 6. 标签热度 ────────────────────────────────
    print("[6/6] 计算标签热度...")
    all_heat = []
    for tag_col, dim_lbl in (
        ("口味_统计", "口味"),
        ("商品主体_统计", "商品主体"),
        ("规格数值", "规格数值"),
        ("价格带", "价格带"),
    ):
        h = calc_tag_heat(df, tag_col, dim_label=dim_lbl)
        if not h.empty:
            all_heat.append(h)
            print(f"      {dim_lbl}：{len(h)} 条")
    heat_df = pd.concat(all_heat, ignore_index=True).reset_index(drop=True)

    # 规格值来源元信息：用于前端展示“最优规格值”的单位与来源主体（按真实来源，不按大类猜测）
    spec_meta = {}
    if all(c in df.columns for c in ["统一大分类", "规格数值", "规格单位", "销量"]):
        _spec_df = df.copy()
        _spec_df["规格数值"] = _spec_df["规格数值"].astype(str).str.strip()
        _spec_df["规格单位"] = _spec_df["规格单位"].astype(str).str.strip()
        _spec_df = _spec_df[
            (_spec_df["规格数值"] != "")
            & (_spec_df["规格数值"].str.lower() != "nan")
        ]
        body_col = "商品主体_统计" if "商品主体_统计" in _spec_df.columns else ("商品主体" if "商品主体" in _spec_df.columns else None)
        for (cat, spec_val), g in _spec_df.groupby(["统一大分类", "规格数值"]):
            gu = g[(g["规格单位"] != "") & (g["规格单位"].str.lower() != "nan")]
            if not gu.empty:
                unit_series = gu.groupby("规格单位")["销量"].sum().sort_values(ascending=False)
                spec_unit = str(unit_series.index[0])
                spec_unit_mixed = len(unit_series.index) > 1
            else:
                spec_unit = ""
                spec_unit_mixed = False
            if body_col:
                gb = g[g[body_col].astype(str).str.strip().str.lower().ne("nan") & g[body_col].astype(str).str.strip().ne("")]
                if not gb.empty:
                    spec_source_body = str(gb.groupby(body_col)["销量"].sum().idxmax())
                else:
                    spec_source_body = ""
            else:
                spec_source_body = ""
            spec_meta[(str(cat), str(spec_val))] = {
                "spec_unit": spec_unit,
                "spec_unit_mixed": bool(spec_unit_mixed),
                "spec_source_body": spec_source_body,
            }

    out_cols = [
        "统一大分类", "标签维度", "标签值", "代表品牌",
        "热度评分", "热度等级", "趋势方向", "机会信号",
        "SKU数量", "覆盖地区数",
        "平均销量PSD", "平均铺货转化率", "平均毛利率", "平均客数PSD", "总销量",
    ]
    heat_out = (
        heat_df[out_cols]
        .sort_values(["统一大分类", "标签维度", "热度评分"], ascending=[True, True, False])
        .reset_index(drop=True)
    )
    # 主体中含“外卖”属于门店操作/场景噪声，不作为商品主体标签展示
    if not heat_out.empty:
        _mask_takeout_body = (
            heat_out["标签维度"].astype(str).eq("商品主体")
            & heat_out["标签值"].astype(str).str.contains("外卖", na=False)
        )
        if _mask_takeout_body.any():
            heat_out = heat_out.loc[~_mask_takeout_body].reset_index(drop=True)

    # ── 输出 Excel ─────────────────────────────────
    print(f"\n输出 Excel → {OUTPUT_FILE}")
    wb = openpyxl.Workbook()

    # Sheet1: 标签热度总览
    ws1 = wb.active; ws1.title = "标签热度总览"
    write_heat_sheet(ws1, heat_out, out_cols)

    # Sheet2: 口味热度
    ws2 = wb.create_sheet("口味热度")
    write_heat_sheet(ws2,
        heat_out[heat_out["标签维度"] == "口味"]
        .sort_values(["统一大分类", "热度评分"], ascending=[True, False]),
        out_cols)

    # Sheet3: 商品主体热度
    ws3 = wb.create_sheet("商品主体热度")
    write_heat_sheet(ws3,
        heat_out[heat_out["标签维度"] == "商品主体"]
        .sort_values(["统一大分类", "热度评分"], ascending=[True, False]),
        out_cols)

    # Sheet4: 规格与价格带热度
    ws4 = wb.create_sheet("规格价格带热度")
    write_heat_sheet(ws4,
        heat_out[heat_out["标签维度"].isin(["规格数值", "价格带"])]
        .sort_values(["统一大分类", "标签维度", "热度评分"], ascending=[True, True, False]),
        out_cols)

    # Sheet5: 商品宽表（下钻用）
    ws5 = wb.create_sheet("商品宽表(下钻用)")
    raw_cols = [
        "渠道来源", "数据期间", "地域名称", "省份", "城市", "渠道类型",
        "大分类名称", "统一大分类",
        "商品名称", "品牌_解析", "商品主体", "商品主体_统计", "口味", "口味_统计",
        "规格数值", "规格单位", "价格带", "包装类型",
        "售价", "生命周期",
        "销量", "客数", "销量PSD", "客数PSD",
        "动销店数", "库存店数", "毛利率", "铺货转化率",
    ]
    df_raw = df[[c for c in raw_cols if c in df.columns]]
    NEW_RAW = {
        "省份", "城市", "渠道类型", "统一大分类",
        "品牌_解析", "商品主体", "商品主体_统计", "口味", "口味_统计",
        "规格数值", "规格单位", "价格带", "包装类型", "铺货转化率",
    }
    for ci, col in enumerate(df_raw.columns, 1):
        hdr(ws5.cell(row=1, column=ci, value=col))
    for ri, row in enumerate(dataframe_to_rows(df_raw, index=False, header=False), 2):
        for ci, val in enumerate(row, 1):
            cn = df_raw.columns[ci - 1]
            cel(ws5.cell(row=ri, column=ci, value=val),
                "new" if cn in NEW_RAW else None)
    for ci, col in enumerate(df_raw.columns, 1):
        ws5.column_dimensions[get_column_letter(ci)].width = \
            {"商品名称": 30, "数据期间": 22, "地域名称": 18}.get(col, 12)
    ws5.freeze_panes = "A2"
    ws5.auto_filter.ref = ws5.dimensions

    # Sheet6: 评分逻辑说明
    ws6 = wb.create_sheet("热度评分说明")
    notes = [
        ("标签热度评分 — 设计理念",                                        True,  "FFFFFF", "1F4E79"),
        ("核心问题：现在哪种口味/类型/规格/价格带正在被消费者接受？",        False, "000000", "E2EFDA"),
        ("计算单位：某标签在某品类下所有商品的整体热度，而非单品评分",        False, "000000", None),
        ("",                                                                False, None,    None),
        ("热度评分 = 当前量级×40% + 趋势方向×35% + 地域广度×25%",         True,  "FFFFFF", "2E75B6"),
        ("当前量级（40%）：销量PSD均值×25% + 铺货转化率×10% + 客数PSD×5%", False, "000000", None),
        ("趋势方向（35%）：跨期销量PSD线性回归，上升→1，下滑→0，单期→0.5", False, "000000", None),
        ("地域广度（25%）：覆盖地区数，越广说明消费者接受度越普遍",          False, "000000", None),
        ("",                                                                False, None,    None),
        ("热度等级",                                                        True,  "FFFFFF", "2E75B6"),
        ("🔥 高热（≥0.75）：量级大且上升，该特征正被消费者广泛接受",        False, "000000", "FCE4D6"),
        ("📈 上升（0.50-0.75）：增长趋势明显，值得持续关注",                False, "000000", "E2EFDA"),
        ("➡️ 平稳（0.25-0.50）：稳定，无明显趋势",                         False, "000000", "EDEDED"),
        ("❄️ 冷淡（<0.25）：接受度低或持续下滑，谨慎引进",                 False, "000000", "FFE0E0"),
        ("⚡ 高需低供：销量PSD超过60分位且SKU数≤2，供给不足，开发机会大",   False, "000000", "FFF2CC"),
        ("",                                                                False, None,    None),
        ("推荐分析路径",                                                    True,  "FFFFFF", "C55A11"),
        ("1. 口味热度      → 发现消费者正在接受的口味方向",                  False, "000000", "FFF2CC"),
        ("2. 商品主体热度  → 确认哪类商品形态在起势",                        False, "000000", "FFF2CC"),
        ("3. 规格/价格带热度 → 锁定目标规格和价格定位",                     False, "000000", "FFF2CC"),
        ("4. 三者交叉      → 高热口味 × 高热主体 × 高热规格 = 新品开发方向", False, "000000", "FFF2CC"),
        ("",                                                                False, None,    None),
        ("注意事项",                                                        True,  "FFFFFF", "C55A11"),
        ("• 评分在当前数据集内归一化，新增数据后需重新运行脚本",             False, "000000", "FFECEC"),
        ("• 趋势方向需要≥2个时间期间的数据，单期商品固定为0.5",              False, "000000", "FFECEC"),
        ("• 口味热度按「口味_统计」聚合：去「味/风味」、苹果系与蜜桃/水蜜桃/白桃等→桃；主体热度按「商品主体_统计」：酸牛奶→酸奶、奶酪益生元→奶酪、食用盐→食盐等", False, "000000", "FFECEC"),
        ("• 补充口味词 → 脚本顶部 FLAVOR_WORDS 列表添加",                   False, "000000", "FFECEC"),
        ("• 补充主体词 → 脚本顶部 PRODUCT_TYPES 列表添加（具体词放前面）",   False, "000000", "FFECEC"),
    ]
    for ri, (text, bold, fc, bg) in enumerate(notes, 1):
        c = ws6.cell(row=ri, column=1, value=text)
        c.font = Font(bold=bold, name="Arial", size=10, color=fc or "000000")
        c.alignment = Alignment(wrap_text=True, vertical="center")
        if bg:
            c.fill = make_fill(bg)
        ws6.row_dimensions[ri].height = 22
    ws6.column_dimensions["A"].width = 90

    wb.save(OUTPUT_FILE)

    # ── 导出 JSON（供前端页面动态读取）────────────────
    import json

    JSON_FILE = os.path.splitext(OUTPUT_FILE)[0] + "_热度数据.json"
    print(f"\n[附加] 导出前端 JSON → {JSON_FILE}")

    def _heat_rows_for(dim_filter=None):
        """将热度 DataFrame 转为前端可用的列表，可按标签维度筛选"""
        sub = heat_out if dim_filter is None else heat_out[heat_out["标签维度"].isin(dim_filter)]
        out = []
        for _, r in sub.iterrows():
            _cat = str(r["统一大分类"])
            _dim = str(r["标签维度"])
            _tag = str(r["标签值"])
            _spec = spec_meta.get((_cat, _tag), {}) if _dim == "规格数值" else {}
            out.append({
                "cat":        _cat,
                "dim":        _dim,
                "tag":        _tag,
                "brand":      str(r["代表品牌"]).strip() if pd.notna(r.get("代表品牌")) and str(r.get("代表品牌", "")).strip() else "",
                "score":      float(r["热度评分"]),
                "level":      r["热度等级"],
                "trend":      r["趋势方向"],
                "signal":     r["机会信号"] if r["机会信号"] else "",
                "sku":        int(r["SKU数量"]),
                "regions":    int(r["覆盖地区数"]),
                "psd":        float(r["平均销量PSD"]),
                "conv":       float(r["平均铺货转化率"]),
                "margin":     float(r["平均毛利率"]),
                "cust_psd":   float(r["平均客数PSD"]),
                "total_qty":  int(r["总销量"]),
                "spec_unit": _spec.get("spec_unit", ""),
                "spec_unit_mixed": _spec.get("spec_unit_mixed", False),
                "spec_source_body": _spec.get("spec_source_body", ""),
            })
        return out

    # 按品类汇总，构造前端所需的完整结构
    cats = sorted(heat_out["统一大分类"].unique().tolist())
    json_payload = {
        "generated_at": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M"),
        "categories": cats,
        "heat": _heat_rows_for(),           # 全量，前端自行筛选
        # 以下为各品类快照数据（侧边栏 TOP 卡片用）
        "snapshots": {}
    }

    for cat in cats:
        sub = heat_out[heat_out["统一大分类"] == cat]

        def top1_by_psd(dim):
            """按销量PSD取第一（口味/主体：反映当前实际销量）"""
            s = sub[sub["标签维度"] == dim].copy()
            if dim == "商品主体":
                # 过滤单期/极窄样本的异常峰值，避免如「单枚茶叶蛋」这类点状样本成为快照 TOP 主体
                s = s[
                    ~s["趋势方向"].astype(str).eq("— 单期")
                    & (s["SKU数量"] >= 2)
                    & (s["覆盖地区数"] >= 5)
                ]
                if s.empty:
                    s = sub[sub["标签维度"] == dim].copy()
            s = s.sort_values("平均销量PSD", ascending=False)
            if s.empty:
                return None
            r = s.iloc[0]
            out = {"tag": r["标签值"], "psd": float(r["平均销量PSD"]),
                   "score": float(r["热度评分"]), "level": r["热度等级"],
                   "trend": r["趋势方向"],
                   "conv": float(r["平均铺货转化率"]),
                   "signal": r["机会信号"] if r["机会信号"] else ""}
            if dim == "商品主体":
                b = str(r.get("代表品牌", "") or "").strip()
                if b:
                    out["brand"] = b
            return out

        def top1_by_score(dim):
            """按热度评分取第一（价格带：反映哪个价格带在增长）"""
            s = sub[sub["标签维度"] == dim].sort_values("热度评分", ascending=False)
            if s.empty:
                return None
            r = s.iloc[0]
            return {"tag": r["标签值"], "psd": float(r["平均销量PSD"]),
                    "score": float(r["热度评分"]), "level": r["热度等级"],
                    "trend": r["趋势方向"],
                    "conv": float(r["平均铺货转化率"]),
                    "total_qty": int(r["总销量"]),
                    "signal": r["机会信号"] if r["机会信号"] else ""}

        # 价格带完整列表（按热度评分排序，前端渲染分布图用）
        price_rows_sorted = sub[sub["标签维度"] == "价格带"].sort_values("热度评分", ascending=False)
        prices_all = [
            {"tag": r["标签值"], "psd": float(r["平均销量PSD"]),
             "score": float(r["热度评分"]), "level": r["热度等级"],
             "trend": r["趋势方向"], "conv": float(r["平均铺货转化率"]),
             "total_qty": int(r["总销量"]), "sku": int(r["SKU数量"])}
            for _, r in price_rows_sorted.iterrows()
        ]

        # 口味列表（侧边栏热度条）
        flavor_rows = (
            sub[sub["标签维度"] == "口味"]
            .sort_values("热度评分", ascending=False)
            .head(10)
        )
        flavors = [
            {"name": r["标签值"], "score": float(r["热度评分"]),
             "signal": r["机会信号"] if r["机会信号"] else ""}
            for _, r in flavor_rows.iterrows()
        ]

        # 空白机会（高需低供）
        blanks = sub[sub["机会信号"] == "⚡ 高需低供"]["标签值"].tolist()

        json_payload["snapshots"][cat] = {
            "top_flavor":    top1_by_psd("口味"),
            "top_body":      top1_by_psd("商品主体"),
            # 价格带双视角：热度增长 vs 当前销量体量
            "top_price":     top1_by_score("价格带"),
            "top_price_vol": top1_by_psd("价格带"),
            "prices_all":    prices_all,
            "price_list":    prices_all,
            "flavors":       flavors,
            "blanks":        blanks,
        }

    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(json_payload, f, ensure_ascii=False, indent=2)
    print(f"      品类数：{len(cats)}，热度条目：{len(json_payload['heat'])}")

    print(f"\n✅ 完成！→ {OUTPUT_FILE}")
    print(f"           → {JSON_FILE}  （前端页面读取此文件）")
    print("\nSheet 说明：")
    print("  标签热度总览    所有标签维度，按品类+维度+热度排序")
    print("  口味热度        仅口味维度，采购人员最常看")
    print("  商品主体热度    仅商品主体维度")
    print("  规格价格带热度  规格数值 + 价格带")
    print("  商品宽表(下钻)  原始数据+解析字段，供具体商品查询")
    print("  热度评分说明    评分逻辑完整说明")
    print("\n词典补充提示：")
    print("  口味识别遗漏 → 在脚本顶部 FLAVOR_WORDS 里添加")
    print("  主体识别遗漏 → 在 PRODUCT_TYPES 里添加（具体词放前面）")


if __name__ == "__main__":
    main()
