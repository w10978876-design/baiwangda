"""
快消品销售数据处理与标签热度分析脚本
=====================================
功能：
  1. 读取两渠道 Excel/TSV 数据，字段对齐、合并
  2. 地域解析（省份 / 城市 / 渠道类型）
  3. 商品名称解析（品牌 / 商品主体 / 口味 / 规格 / 包装）
  4. 统一大分类映射
  5. 标签热度计算（口味 / 商品主体 / 规格 / 价格带）
     热度 = 当前量级×40% + 时间趋势×35% + 地域广度×25%
  6. 输出 Excel 报告（标签热度总览 + 商品宽表）

使用方法：
  1. 修改下方 ══ 配置区 ══ 里的文件路径
  2. 运行：python fmcg_analysis.py
  3. 查看输出文件

依赖：pip install pandas openpyxl numpy
"""

import re
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


# ══════════════════════════════════════════════════
# 配置区 — 根据实际情况修改这里
# ══════════════════════════════════════════════════

# 输入文件（支持 .xlsx / .xls / .txt / .tsv / .csv）
# 注意：必须与 CAT_MAP 的键一致——「渠道A」对应壹度表的大分类命名（需 CHANNEL_A_RENAME），
# 「渠道B」对应安达表的大分类命名。若两路文件与脚本标注对调，映射大量失效，页面会像「两套分类」重复出现。
INPUT_FILES = {
    "渠道A": "data/壹度2512-2601-2603-2503-2504.xlsx",
    "渠道B": "data/安达2512-2601-2603-2503-2504.xlsx",
}

# 渠道A的字段重命名（渠道A字段名与标准不同）
CHANNEL_A_RENAME = {
    "时间期间": "数据期间",
    "商品状态": "生命周期",
}

# 输出文件路径
OUTPUT_FILE = "商品标签热度分析.xlsx"

# 文件读取分隔符（TSV用'\t'，CSV用','，Excel文件改成 None）
FILE_SEP = None


# ══════════════════════════════════════════════════
# 词典区 — 根据业务补充和调整
# ══════════════════════════════════════════════════

# 统一大分类映射：(渠道来源, 原始大分类名) → 统一大分类
CAT_MAP = {
    ("渠道A", "FF类"):          "FF速食",
    ("渠道A", "休闲食品"):      "休闲零食",
    ("渠道A", "槟榔"):          "休闲零食",
    ("渠道A", "糖果巧克力"):    "休闲零食",
    ("渠道A", "饼干膨化"):      "休闲零食",
    ("渠道A", "低温日配"):      "日配冷藏",
    ("渠道A", "冲调方便速食"):  "方便速食",
    ("渠道A", "冷冻食品"):      "冷冻食品",
    ("渠道A", "常温奶"):        "常温乳品",
    ("渠道A", "常温饮料"):      "饮料",
    ("渠道A", "烘焙面包"):      "日配烘焙",
    ("渠道A", "粮油调味"):      "粮油调味",
    ("渠道A", "酒类"):          "酒类",
    ("渠道A", "日化美护"):      "日化美护",
    ("渠道A", "计生医疗"):      "日化美护",
    ("渠道A", "家居百货"):      "家居百货",

    ("渠道B", "加热速食"):      "FF速食",
    ("渠道B", "休闲零食"):      "休闲零食",
    ("渠道B", "饼干糖巧"):      "休闲零食",
    ("渠道B", "低温速食"):      "日配冷藏",
    ("渠道B", "日配冷藏"):      "日配冷藏",
    ("渠道B", "方便食品"):      "方便速食",
    ("渠道B", "冷冻"):          "冷冻食品",
    ("渠道B", "常温乳品"):      "常温乳品",
    ("渠道B", "饮料"):          "饮料",
    ("渠道B", "日配烘焙"):      "日配烘焙",
    ("渠道B", "粮油副食"):      "粮油调味",
    ("渠道B", "酒"):            "酒类",
    ("渠道B", "日化美护"):      "日化美护",
    ("渠道B", "家居百货"):      "家居百货",
}

# 已知城市白名单（防止误识别，可继续添加）
KNOWN_CITIES = [
    "包头", "呼和浩特", "赤峰", "通辽", "鄂尔多斯",
    "中牟", "濮阳", "郑州", "开封", "洛阳", "安阳", "焦作",
    "新乡", "许昌", "漯河", "南阳", "商丘", "信阳", "周口", "驻马店",
    "北京", "上海", "广州", "深圳", "成都", "武汉", "西安",
    "杭州", "南京", "天津", "重庆", "沈阳", "长春", "哈尔滨",
    "济南", "青岛", "合肥", "福州", "厦门", "南昌",
    "长沙", "昆明", "贵阳", "南宁", "海口",
    "太原", "石家庄", "兰州", "西宁", "银川", "乌鲁木齐",
]

# 渠道类型关键词映射
CHANNEL_MAP = {
    "便利":   "便利店",
    "超市":   "超市",
    "大卖场": "大卖场",
    "社区":   "社区店",
    "线上":   "线上",
    "电商":   "线上",
    "KA":     "KA卖场",
    "分公司": "分公司门店",
}

# 省份列表
PROVINCES = [
    "北京", "天津", "上海", "重庆",
    "河北", "山西", "辽宁", "吉林", "黑龙江",
    "江苏", "浙江", "安徽", "福建", "江西", "山东",
    "河南", "湖北", "湖南", "广东", "海南",
    "四川", "贵州", "云南", "陕西", "甘肃", "青海",
    "内蒙古", "广西", "西藏", "宁夏", "新疆",
]

# 口味词典（解析时按字符串长度从长到短匹配，复合/长尾词请写完整如「青柠味」）
FLAVOR_WORDS = [
    "日式浇汁", "蜂蜜芥末", "火鸡面风味", "黑胡椒", "奥尔良",
    "照烧", "香辣", "麻辣", "红烧", "咖喱", "蒜香", "酱香",
    "香菇", "糯米", "蟹味", "原味", "清淡", "甜味", "咸味",
    "番茄", "草莓", "芒果", "抹茶", "黑糖", "海鲜", "芝士",
    "浸汁", "炭烤",
    "青柠味", "柠檬味", "白桃味", "水蜜桃味", "葡萄味", "柑橘味", "西柚味",
    "葡萄伏特加风味", "热带风味", "血橙味",
    "荔枝海盐味", "海盐荔枝味", "荔枝味",
    "水蜜桃", "柠檬", "青柠", "白桃", "荔枝",
    # ↑ 在这里继续补充遗漏的口味词
]

FLAVOR_WORDS_SORTED = sorted(set(FLAVOR_WORDS), key=len, reverse=True)

# 商品名开头的已知品牌/子品牌（越长越优先匹配，用于无字段或字段明显错误时纠偏）
KNOWN_BRAND_PREFIXES = sorted(
    {
        "农夫山泉东方树叶",
        "农夫山泉水溶C",
        "元气森林外星人",
        "元气森林",
        "农夫山泉",
        "东方树叶",
        "水溶C",
        "江小白果立方",
        "江小白",
        "凤城老窖",
        "海底捞",
        "双汇",
        "壹度定制",
        "五得利",
        "红星",
        "汾酒",
        "郎酒",
        "锐澳",
        "娃哈哈",
        "康师傅",
        "统一企业",
        "统一",
        "怡宝",
        "百岁山",
        "名仁",
        "可口可乐",
        "百事",
        "雪碧",
        "芬达",
        "东鹏",
        "红牛",
        "脉动",
        "外星人",  # 名单独出现时作品牌前缀（若前面已有元气森林会在更长前缀中吃掉）
    },
    key=len,
    reverse=True,
)

# ERP「商品品牌」明显占位/错误时，以商品名称前缀推断为准
BRAND_FIELD_BLACKLIST = frozenset({"圣诞", "元旦", "测试", "未知", "无", "—", "-"})

# 包装类型词典
PACKAGE_WORDS = ["罐装", "瓶装", "袋装", "盒装", "礼盒", "散装", "桶装", "箱装"]

# 商品主体词典（具体词放前面优先匹配，可继续补充）
PRODUCT_TYPES = [
    # FF速食 ─────────────────
    (r"鸡肉包",              "鸡肉包"),
    (r"牛肉包",              "牛肉包"),
    (r"油菜包",              "油菜包"),
    (r"肉包",                "肉包"),
    (r"烧麦",                "烧麦"),
    (r"鸡肉饭团|饭团",       "鸡肉饭团"),
    (r"地道肠",              "地道肠"),
    (r"鸡肉肠|开花鸡肉肠",  "鸡肉肠"),
    (r"鸡肉串",              "鸡肉串"),
    (r"腿肉串",              "腿肉串"),
    (r"猪肉串",              "猪肉串"),
    (r"唐扬串",              "唐扬串"),
    (r"炭烤鸡全腿|鸡全腿",  "鸡全腿"),
    (r"鸡丸串|丸串",         "鸡丸串"),
    (r"海带串",              "海带串"),
    (r"面筋",                "面筋"),
    (r"卤蛋",                "卤蛋"),
    (r"蛋挞",                "蛋挞"),
    (r"玉米",                "玉米"),
    (r"豆奶|豆浆",           "豆奶"),
    # 酒类 ───────────────────
    (r"二锅头",              "二锅头"),
    (r"高粱酒",              "高粱酒"),
    (r"白酒",                "白酒"),
    (r"啤酒",                "啤酒"),
    (r"葡萄酒|红酒",         "葡萄酒"),
    (r"鸡尾酒|强爽",          "鸡尾酒"),
    (r"果汁酒",               "果汁酒"),
    (r"地瓜烧",               "地瓜烧"),
    (r"茉莉花茶",             "茉莉花茶"),
    (r"冰红茶饮料|冰红茶",     "冰红茶饮料"),
    (r"复合果汁",             "复合果汁"),
    (r"午餐肉罐头|午餐肉",     "午餐肉罐头"),
    (r"白砂糖",               "白砂糖"),
    (r"小麦粉",               "小麦粉"),
    (r"火锅底料",             "火锅底料"),
    # 饮料水饮（具体词放前，避免「电解质」早于「电解质水」被截断）────────
    (r"电解质水|电解制水",    "电解质水"),
    (r"^电解质$|^电解$",      "电解质水"),
    (r"苏打水|苏打气泡水",    "苏打水"),
    (r"纯净水|蒸馏水|纯水",    "纯净水"),
    (r"天然矿泉水|矿泉水|天然水", "矿泉水"),
    (r"气泡水|含气水",        "气泡水"),
    (r"饮用水|直饮水",        "饮用水"),
    # ↑ 在这里继续补充遗漏的主体词（具体词放前面）
]

# 去掉口味/规格后剩余短词 → 规范主体（如「电解质」→「电解质水」）
BODY_REMAINDER_MAP = {
    "电解质": "电解质水",
    "电解": "电解质水",
    "纯水": "纯净水",
    "矿泉水": "矿泉水",
    "苏打": "苏打水",
}

# 规格正则（容量 / 重量，与度数严格区分）
SPEC_RE = re.compile(
    r"(\d+(?:\.\d+)?)\s*"
    r"(ml|ML|mL|L|升|g|G|kg|KG|克|斤|包|瓶|罐|袋|片|粒|支|盒|桶|箱|个|串|根|条|只)"
)

# 度数正则（酒类特有，归入口味字段）
DEGREE_RE = re.compile(r"(\d+(?:\.\d+)?)\s*[°度]")

# 价格带分箱
PRICE_BINS   = [0, 3, 6, 10, 20, 9999]
PRICE_LABELS = ["低价(≤3元)", "亲民(3-6元)", "中价(6-10元)", "中高(10-20元)", "高价(20元+)"]


# ══════════════════════════════════════════════════
# 解析函数
# ══════════════════════════════════════════════════

def read_file(path, source_name, sep="\t"):
    """读取数据文件，自动识别 Excel 与文本格式"""
    if sep is None or path.endswith((".xlsx", ".xls")):
        df = pd.read_excel(path)
    else:
        df = pd.read_csv(path, sep=sep)
    df["渠道来源"] = source_name
    return df


def parse_region(name):
    """
    解析地域名称 → 省份 / 城市 / 渠道类型

    格式1：'中牟-张志鹏'       → 城市=中牟（连字符前为城市）
    格式2：'安哒便利包头分公司' → 城市=包头（白名单扫描）
    """
    name = str(name).strip()
    province = city = channel = ""

    for p in PROVINCES:
        if p in name:
            province = p
            break

    m = re.match(r"^([^\-\—]+)[\-\—]", name)
    if m:
        city = m.group(1).strip()
    if not city or city not in KNOWN_CITIES:
        for c in KNOWN_CITIES:
            if c in name:
                city = c
                break

    for kw, ch in CHANNEL_MAP.items():
        if kw in name:
            channel = ch
            break

    return pd.Series([province, city, channel])


def _infer_brand_prefix(name: str) -> str:
    """取商品名最前缀的已知品牌（长词优先）。"""
    n = str(name).strip()
    # 某些源数据会带单字符前缀（如 z元气森林...），先清掉前导杂质再识别品牌
    n = re.sub(r"^[^一-龥A-Za-z0-9]+", "", n)
    n = re.sub(r"^[A-Za-z]{1,3}\s*", "", n)
    for p in KNOWN_BRAND_PREFIXES:
        if n.startswith(p):
            return p
    return ""


def _resolve_brand(raw: str, brand_field: str) -> str:
    """
    品牌：优先与商品名前缀一致的「商品品牌」字段；字段缺失、黑名单或与名称矛盾时，
    用名称前缀的已知品牌纠偏（如 名仁水蜜桃苏打水 纠正 ERP「圣诞」）。
    """
    b = re.sub(r"\[.*?\]", "", str(brand_field)).strip()
    if b in ("", "-", "未知", "nan"):
        b = ""
    inferred = _infer_brand_prefix(raw)
    if b and b in BRAND_FIELD_BLACKLIST:
        return inferred or b
    raw_norm = re.sub(r"^[^一-龥A-Za-z0-9]+", "", raw)
    raw_norm = re.sub(r"^[A-Za-z]{1,3}\s*", "", raw_norm)
    # 子品牌归母品牌+子品牌，保证展示稳定
    if raw_norm.startswith("东方树叶"):
        return "农夫山泉东方树叶"
    if raw_norm.startswith("水溶C"):
        return "农夫山泉水溶C"
    if raw_norm.startswith("顺品郎"):
        return "郎酒"
    if b and raw_norm.startswith(b):
        return b
    if inferred:
        return inferred
    return b


def parse_product(name, brand_field=""):
    """
    解析商品名称 → 品牌_解析 / 商品主体 / 口味 / 规格数值 / 规格单位 / 包装类型

    解析逻辑：
      品牌：字段与名称前缀一致则用字段；否则用已知品牌前缀（含「元气森林外星人」）
      度数（酒类）→ 归入口味字段（不是规格）
      口味：按词长优先匹配 FLAVOR_WORDS
      商品主体：PRODUCT_TYPES 优先；饮料类如 纯净水/电解质水/苏打水；余字映射 BODY_REMAINDER_MAP
    """
    raw = str(name).strip()
    raw_norm = re.sub(r"^[^一-龥A-Za-z0-9]+", "", raw)
    raw_norm = re.sub(r"^[A-Za-z]{1,3}\s*", "", raw_norm)
    brand = _resolve_brand(raw, brand_field)

    # 1. 提取度数
    degree = ""
    dm = DEGREE_RE.search(raw_norm)
    if dm:
        degree = dm.group(1) + "°"

    # 2. 去度数 → 去品牌前缀，得到待拆分的中间串
    work = DEGREE_RE.sub("", raw_norm).strip()
    if brand:
        work = re.sub(r"^" + re.escape(brand) + r"\s*", "", work).strip()

    # 3. 口味（长词优先；无词时用度数）
    flavor = ""
    for f in FLAVOR_WORDS_SORTED:
        if f in work:
            flavor = f
            break
    if not flavor and degree:
        flavor = degree

    # 4. 规格（容量/重量）
    spec_val = spec_unit = ""
    m = SPEC_RE.search(raw)
    if m:
        spec_val, spec_unit = m.group(1), m.group(2)

    # 5. 包装类型
    package = ""
    for p in PACKAGE_WORDS:
        if p in raw:
            package = p
            break

    # 6. 商品主体：先去掉口味（非度数）再匹配类型词
    work_for_type = work
    if flavor and flavor != degree:
        work_for_type = work.replace(flavor, "", 1).strip()

    product_type = ""
    for pattern, label in PRODUCT_TYPES:
        if re.search(pattern, work_for_type):
            product_type = label
            break

    if product_type:
        body = product_type
    else:
        body = work_for_type
        body = SPEC_RE.sub("", body)
        if flavor and flavor != degree:
            body = body.replace(flavor, "")
        for p in PACKAGE_WORDS:
            body = body.replace(p, "")
        # 仅去掉末尾明显货号式英文，避免误伤中文主体
        body = re.sub(r"[A-Za-z0-9]{3,}$", "", body)
        body = re.sub(r"[（(][^）)]*[）)]", "", body)
        body = re.sub(r"[\s\-\_\.]+", "", body).strip()
        if body in BODY_REMAINDER_MAP:
            body = BODY_REMAINDER_MAP[body]
        if len(body) <= 1:
            body = brand if brand else raw

    return pd.Series([brand, body, flavor, spec_val, spec_unit, package])


# ══════════════════════════════════════════════════
# 热度计算
# ══════════════════════════════════════════════════

def calc_tag_heat(df, tag_col):
    """
    计算标签维度热度

    热度评分 = 当前量级×40% + 时间趋势×35% + 地域广度×25%
      当前量级：销量PSD(25%) + 铺货转化率(10%) + 客数PSD(5%)
      时间趋势：同标签跨期销量PSD线性回归斜率（单期→0.5，不奖不罚）
      地域广度：该标签覆盖的地区数

    评分在每个(品类×标签维度)内归一化，仅用于同类横向比较。
    """
    valid = df[
        df[tag_col].notna() &
        (df[tag_col].astype(str).str.strip() != "") &
        (df[tag_col].astype(str) != "nan")
    ].copy()
    if valid.empty:
        return pd.DataFrame()

    rows = []
    for (cat, tag), g in valid.groupby(["统一大分类", tag_col]):

        # 时间趋势（按库存店数加权的期间PSD均值，跨期线性回归）
        g2 = g.copy()
        g2["_w"] = np.where(g2["库存店数"] > 0, g2["库存店数"], 1)
        g2["_wpsd"] = g2["销量PSD"] * g2["_w"]
        by_period = (
            g2.groupby("数据期间", as_index=False)[["_wpsd", "_w"]].sum()
            .assign(期PSD=lambda x: x["_wpsd"] / x["_w"])
            .sort_values("数据期间")
        )
        n_p = len(by_period)
        if n_p >= 2:
            x = np.arange(n_p, dtype=float)
            y = by_period["期PSD"].values.astype(float)
            mean_y = y.mean() if y.mean() != 0 else 1.0
            slope = np.polyfit(x, y, 1)[0] / mean_y
            trend_score = float(np.clip((slope + 1) / 2, 0, 1))
            trend_label = ("↑ 上升" if slope > 0.1 else
                           "↓ 下滑" if slope < -0.1 else "→ 平稳")
        else:
            trend_score = 0.5
            trend_label = "— 单期"

        rep_brand = ""
        if tag_col == "商品主体" and "品牌_解析" in g.columns and "销量" in g.columns:
            gb = g.copy()
            gb["品牌_解析"] = gb["品牌_解析"].astype(str).str.strip()
            gb = gb[gb["品牌_解析"].notna() & (gb["品牌_解析"] != "") & (gb["品牌_解析"] != "nan")]
            if not gb.empty:
                rep_brand = str(gb.groupby("品牌_解析")["销量"].sum().idxmax())

        rows.append({
            "统一大分类":     cat,
            "标签维度":       tag_col,
            "标签值":         str(tag),
            "代表品牌":       rep_brand,
            "SKU数量":        g["商品名称"].nunique(),
            "覆盖地区数":     g["地域名称"].nunique(),
            "平均销量PSD":    round(g["销量PSD"].mean(), 3),
            "平均铺货转化率": round(g["铺货转化率"].mean(), 3),
            "平均毛利率":     round(g["毛利率"].mean(), 3),
            "平均客数PSD":    round(g["客数PSD"].mean(), 3),
            "总销量":         int(g["销量"].sum()),
            "趋势方向":       trend_label,
            "_t":  trend_score,
            "_p":  g["销量PSD"].mean(),
            "_c":  g["铺货转化率"].mean(),
            "_cl": g["客数PSD"].mean(),
            "_r":  float(g["地域名称"].nunique()),
        })

    result = pd.DataFrame(rows)

    # (品类×维度)内归一化，用 transform 避免 groupby 丢列
    for src, dst in [("_p","_np"), ("_c","_nc"), ("_cl","_ncl"), ("_t","_nt"), ("_r","_nr")]:
        result[dst] = result.groupby(["统一大分类", "标签维度"])[src].transform(
            lambda s: ((s - s.min()) / (s.max() - s.min())).clip(0, 1)
            if s.max() > s.min() else 0.5
        )

    result["热度评分"] = (
        (result["_np"] * 0.25 + result["_nc"] * 0.10 + result["_ncl"] * 0.05) * 0.40
        + result["_nt"] * 0.35
        + result["_nr"] * 0.25
    ).round(4)

    result["热度等级"] = result["热度评分"].apply(
        lambda s: "🔥 高热" if s >= 0.75 else
                  "📈 上升" if s >= 0.50 else
                  "➡️ 平稳" if s >= 0.25 else "❄️ 冷淡"
    )

    # 机会信号：销量PSD高但SKU少 = 供给不足，开发机会
    psd_q60 = result["平均销量PSD"].quantile(0.6)
    result["机会信号"] = result.apply(
        lambda r: "⚡ 高需低供" if (r["平均销量PSD"] > psd_q60 and r["SKU数量"] <= 2)
                  else "", axis=1
    )

    return result


# ══════════════════════════════════════════════════
# Excel 样式
# ══════════════════════════════════════════════════

def make_fill(hex6):
    return PatternFill("solid", start_color=hex6)

FILLS = {
    "header": make_fill("1F4E79"),
    "new":    make_fill("E2EFDA"),
    "hot":    make_fill("FCE4D6"),
    "rise":   make_fill("E2EFDA"),
    "flat":   make_fill("EDEDED"),
    "fall":   make_fill("FFE0E0"),
    "stripe": make_fill("EBF3FB"),
    "score":  make_fill("FFF2CC"),
}
H_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
N_FONT = Font(name="Arial", size=10)
B_FONT = Font(name="Arial", size=10, bold=True)
C_ALIGN = Alignment(horizontal="center", vertical="center")
L_ALIGN = Alignment(horizontal="left",   vertical="center")
THIN = Border(**{s: Side(style="thin", color="CCCCCC")
                 for s in ("left", "right", "top", "bottom")})

def hdr(cell, key="header"):
    cell.font = H_FONT; cell.fill = FILLS[key]
    cell.alignment = C_ALIGN; cell.border = THIN

def cel(cell, key=None, bold=False):
    cell.font = B_FONT if bold else N_FONT
    cell.border = THIN; cell.alignment = L_ALIGN
    if key: cell.fill = FILLS[key]

def write_heat_sheet(ws, data, out_cols):
    col_w = {
        "统一大分类": 12, "标签维度": 10, "标签值": 16, "代表品牌": 12,
        "热度评分": 10, "热度等级": 10, "趋势方向": 10, "机会信号": 12,
        "SKU数量": 8, "覆盖地区数": 10, "平均销量PSD": 12,
        "平均铺货转化率": 12, "平均毛利率": 10, "平均客数PSD": 10, "总销量": 10,
    }
    for ci, col in enumerate(out_cols, 1):
        hdr(ws.cell(row=1, column=ci, value=col))
    prev_cat = prev_dim = None
    score_idx = out_cols.index("热度评分")
    for ri, row in enumerate(dataframe_to_rows(data, index=False, header=False), 2):
        cur_cat, cur_dim = row[0], row[1]
        score = row[score_idx]
        for ci, val in enumerate(row, 1):
            cn = out_cols[ci - 1]
            if cn == "热度评分":
                k = ("hot"  if score >= 0.75 else
                     "rise" if score >= 0.50 else
                     "flat" if score >= 0.25 else "fall")
            elif cur_cat != prev_cat or cur_dim != prev_dim:
                k = "stripe"
            else:
                k = None
            cel(ws.cell(row=ri, column=ci, value=val), k, bold=(cn == "热度评分"))
        prev_cat = cur_cat; prev_dim = cur_dim
    for ci, col in enumerate(out_cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_w.get(col, 12)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ══════════════════════════════════════════════════
# 主流程
# ══════════════════════════════════════════════════

def main():
    print("=" * 55)
    print("  快消品标签热度分析脚本")
    print("=" * 55)

    # ── 1. 读取 ────────────────────────────────────
    print("\n[1/6] 读取数据...")
    dfs = []
    for source, path in INPUT_FILES.items():
        df_tmp = read_file(path, source, sep=FILE_SEP)
        if source == "渠道A":
            df_tmp = df_tmp.rename(columns=CHANNEL_A_RENAME)
        dfs.append(df_tmp)
        print(f"      {source}: {len(df_tmp):,} 行  ←  {path}")
    df = pd.concat(dfs, ignore_index=True)
    print(f"      合并后共 {len(df):,} 行")

    # ── 2. 分类映射 ────────────────────────────────
    print("\n[2/6] 统一大分类映射...")
    df["大分类名称"] = df["大分类名称"].astype(str).str.strip()
    df["统一大分类"] = df.apply(
        lambda r: CAT_MAP.get((r["渠道来源"], r["大分类名称"]), r["大分类名称"]),
        axis=1
    )

    _n_cat = len(df)
    df = df[df["统一大分类"] != "现制品"].copy()
    if len(df) < _n_cat:
        print(f"      已剔除「现制品」{_n_cat - len(df):,} 行")

    # ── 3. 地域解析 ────────────────────────────────
    print("[3/6] 地域解析...")
    df[["省份", "城市", "渠道类型"]] = df["地域名称"].apply(parse_region)

    # ── 4. 商品名称解析 ────────────────────────────
    print("[4/6] 商品名称解析（数据量大时耗时较长）...")
    df[["品牌_解析", "商品主体", "口味", "规格数值", "规格单位", "包装类型"]] = df.apply(
        lambda r: parse_product(r["商品名称"], r.get("商品品牌", "")),
        axis=1
    )

    # ── 5. 数值清洗 ────────────────────────────────
    print("[5/6] 数值清洗与指标计算...")
    for col in ["销量PSD", "销售额PSD", "客数PSD",
                "动销店数", "库存店数", "毛利率",
                "销量", "销售额", "客数", "售价"]:
        if col in df.columns:
            df[col] = pd.to_numeric(
                df[col].astype(str).str.replace("%", "").str.strip(),
                errors="coerce"
            ).fillna(0)

    # 毛利率可能混合两种口径：0-1 与 0-100，按行归一化
    df["毛利率"] = df["毛利率"].apply(lambda x: x / 100 if x > 1 else x)

    df["铺货转化率"] = df.apply(
        lambda r: round(r["动销店数"] / r["库存店数"], 4) if r["库存店数"] > 0 else 0,
        axis=1
    )

    df["价格带"] = pd.cut(
        df["售价"], bins=PRICE_BINS, labels=PRICE_LABELS, right=True
    ).astype(str).replace("nan", "未知价格")

    # ── 6. 标签热度 ────────────────────────────────
    print("[6/6] 计算标签热度...")
    all_heat = []
    for dim in ["口味", "商品主体", "规格数值", "价格带"]:
        h = calc_tag_heat(df, dim)
        if not h.empty:
            all_heat.append(h)
            print(f"      {dim}：{len(h)} 条")
    heat_df = pd.concat(all_heat, ignore_index=True).reset_index(drop=True)

    out_cols = [
        "统一大分类", "标签维度", "标签值", "代表品牌",
        "热度评分", "热度等级", "趋势方向", "机会信号",
        "SKU数量", "覆盖地区数",
        "平均销量PSD", "平均铺货转化率", "平均毛利率", "平均客数PSD", "总销量",
    ]
    heat_out = (
        heat_df[out_cols]
        .sort_values(["统一大分类", "标签维度", "热度评分"], ascending=[True, True, False])
        .reset_index(drop=True)
    )

    # ── 输出 Excel ─────────────────────────────────
    print(f"\n输出 Excel → {OUTPUT_FILE}")
    wb = openpyxl.Workbook()

    # Sheet1: 标签热度总览
    ws1 = wb.active; ws1.title = "标签热度总览"
    write_heat_sheet(ws1, heat_out, out_cols)

    # Sheet2: 口味热度
    ws2 = wb.create_sheet("口味热度")
    write_heat_sheet(ws2,
        heat_out[heat_out["标签维度"] == "口味"]
        .sort_values(["统一大分类", "热度评分"], ascending=[True, False]),
        out_cols)

    # Sheet3: 商品主体热度
    ws3 = wb.create_sheet("商品主体热度")
    write_heat_sheet(ws3,
        heat_out[heat_out["标签维度"] == "商品主体"]
        .sort_values(["统一大分类", "热度评分"], ascending=[True, False]),
        out_cols)

    # Sheet4: 规格与价格带热度
    ws4 = wb.create_sheet("规格价格带热度")
    write_heat_sheet(ws4,
        heat_out[heat_out["标签维度"].isin(["规格数值", "价格带"])]
        .sort_values(["统一大分类", "标签维度", "热度评分"], ascending=[True, True, False]),
        out_cols)

    # Sheet5: 商品宽表（下钻用）
    ws5 = wb.create_sheet("商品宽表(下钻用)")
    raw_cols = [
        "渠道来源", "数据期间", "地域名称", "省份", "城市", "渠道类型",
        "大分类名称", "统一大分类",
        "商品名称", "品牌_解析", "商品主体", "口味",
        "规格数值", "规格单位", "价格带", "包装类型",
        "售价", "生命周期",
        "销量", "客数", "销量PSD", "客数PSD",
        "动销店数", "库存店数", "毛利率", "铺货转化率",
    ]
    df_raw = df[[c for c in raw_cols if c in df.columns]]
    NEW_RAW = {
        "省份", "城市", "渠道类型", "统一大分类",
        "品牌_解析", "商品主体", "口味",
        "规格数值", "规格单位", "价格带", "包装类型", "铺货转化率",
    }
    for ci, col in enumerate(df_raw.columns, 1):
        hdr(ws5.cell(row=1, column=ci, value=col))
    for ri, row in enumerate(dataframe_to_rows(df_raw, index=False, header=False), 2):
        for ci, val in enumerate(row, 1):
            cn = df_raw.columns[ci - 1]
            cel(ws5.cell(row=ri, column=ci, value=val),
                "new" if cn in NEW_RAW else None)
    for ci, col in enumerate(df_raw.columns, 1):
        ws5.column_dimensions[get_column_letter(ci)].width = \
            {"商品名称": 30, "数据期间": 22, "地域名称": 18}.get(col, 12)
    ws5.freeze_panes = "A2"
    ws5.auto_filter.ref = ws5.dimensions

    # Sheet6: 评分逻辑说明
    ws6 = wb.create_sheet("热度评分说明")
    notes = [
        ("标签热度评分 — 设计理念",                                        True,  "FFFFFF", "1F4E79"),
        ("核心问题：现在哪种口味/类型/规格/价格带正在被消费者接受？",        False, "000000", "E2EFDA"),
        ("计算单位：某标签在某品类下所有商品的整体热度，而非单品评分",        False, "000000", None),
        ("",                                                                False, None,    None),
        ("热度评分 = 当前量级×40% + 趋势方向×35% + 地域广度×25%",         True,  "FFFFFF", "2E75B6"),
        ("当前量级（40%）：销量PSD均值×25% + 铺货转化率×10% + 客数PSD×5%", False, "000000", None),
        ("趋势方向（35%）：跨期销量PSD线性回归，上升→1，下滑→0，单期→0.5", False, "000000", None),
        ("地域广度（25%）：覆盖地区数，越广说明消费者接受度越普遍",          False, "000000", None),
        ("",                                                                False, None,    None),
        ("热度等级",                                                        True,  "FFFFFF", "2E75B6"),
        ("🔥 高热（≥0.75）：量级大且上升，该特征正被消费者广泛接受",        False, "000000", "FCE4D6"),
        ("📈 上升（0.50-0.75）：增长趋势明显，值得持续关注",                False, "000000", "E2EFDA"),
        ("➡️ 平稳（0.25-0.50）：稳定，无明显趋势",                         False, "000000", "EDEDED"),
        ("❄️ 冷淡（<0.25）：接受度低或持续下滑，谨慎引进",                 False, "000000", "FFE0E0"),
        ("⚡ 高需低供：销量PSD超过60分位且SKU数≤2，供给不足，开发机会大",   False, "000000", "FFF2CC"),
        ("",                                                                False, None,    None),
        ("推荐分析路径",                                                    True,  "FFFFFF", "C55A11"),
        ("1. 口味热度      → 发现消费者正在接受的口味方向",                  False, "000000", "FFF2CC"),
        ("2. 商品主体热度  → 确认哪类商品形态在起势",                        False, "000000", "FFF2CC"),
        ("3. 规格/价格带热度 → 锁定目标规格和价格定位",                     False, "000000", "FFF2CC"),
        ("4. 三者交叉      → 高热口味 × 高热主体 × 高热规格 = 新品开发方向", False, "000000", "FFF2CC"),
        ("",                                                                False, None,    None),
        ("注意事项",                                                        True,  "FFFFFF", "C55A11"),
        ("• 评分在当前数据集内归一化，新增数据后需重新运行脚本",             False, "000000", "FFECEC"),
        ("• 趋势方向需要≥2个时间期间的数据，单期商品固定为0.5",              False, "000000", "FFECEC"),
        ("• 补充口味词 → 脚本顶部 FLAVOR_WORDS 列表添加",                   False, "000000", "FFECEC"),
        ("• 补充主体词 → 脚本顶部 PRODUCT_TYPES 列表添加（具体词放前面）",   False, "000000", "FFECEC"),
    ]
    for ri, (text, bold, fc, bg) in enumerate(notes, 1):
        c = ws6.cell(row=ri, column=1, value=text)
        c.font = Font(bold=bold, name="Arial", size=10, color=fc or "000000")
        c.alignment = Alignment(wrap_text=True, vertical="center")
        if bg:
            c.fill = make_fill(bg)
        ws6.row_dimensions[ri].height = 22
    ws6.column_dimensions["A"].width = 90

    wb.save(OUTPUT_FILE)

    # ── 导出 JSON（供前端页面动态读取）────────────────
    import json, os

    JSON_FILE = os.path.splitext(OUTPUT_FILE)[0] + "_热度数据.json"
    print(f"\n[附加] 导出前端 JSON → {JSON_FILE}")

    def _heat_rows_for(dim_filter=None):
        """将热度 DataFrame 转为前端可用的列表，可按标签维度筛选"""
        sub = heat_out if dim_filter is None else heat_out[heat_out["标签维度"].isin(dim_filter)]
        out = []
        for _, r in sub.iterrows():
            out.append({
                "cat":        r["统一大分类"],
                "dim":        r["标签维度"],
                "tag":        r["标签值"],
                "brand":      str(r["代表品牌"]).strip() if pd.notna(r.get("代表品牌")) and str(r.get("代表品牌", "")).strip() else "",
                "score":      float(r["热度评分"]),
                "level":      r["热度等级"],
                "trend":      r["趋势方向"],
                "signal":     r["机会信号"] if r["机会信号"] else "",
                "sku":        int(r["SKU数量"]),
                "regions":    int(r["覆盖地区数"]),
                "psd":        float(r["平均销量PSD"]),
                "conv":       float(r["平均铺货转化率"]),
                "margin":     float(r["平均毛利率"]),
                "cust_psd":   float(r["平均客数PSD"]),
                "total_qty":  int(r["总销量"]),
            })
        return out

    # 按品类汇总，构造前端所需的完整结构
    cats = sorted(heat_out["统一大分类"].unique().tolist())
    json_payload = {
        "generated_at": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M"),
        "categories": cats,
        "heat": _heat_rows_for(),           # 全量，前端自行筛选
        # 以下为各品类快照数据（侧边栏 TOP 卡片用）
        "snapshots": {}
    }

    for cat in cats:
        sub = heat_out[heat_out["统一大分类"] == cat]

        def top1(dim, metric="psd"):
            s = sub[sub["标签维度"] == dim].sort_values("平均销量PSD", ascending=False)
            if s.empty:
                return None
            r = s.iloc[0]
            out = {"tag": r["标签值"], "psd": float(r["平均销量PSD"]),
                   "conv": float(r["平均铺货转化率"]), "signal": r["机会信号"] if r["机会信号"] else ""}
            if dim == "商品主体":
                b = str(r.get("代表品牌", "") or "").strip()
                if b:
                    out["brand"] = b
            return out

        # 口味列表（侧边栏热度条）
        flavor_rows = (
            sub[sub["标签维度"] == "口味"]
            .sort_values("热度评分", ascending=False)
            .head(10)
        )
        flavors = [
            {"name": r["标签值"], "score": float(r["热度评分"]),
             "signal": r["机会信号"] if r["机会信号"] else ""}
            for _, r in flavor_rows.iterrows()
        ]

        # 空白机会（高需低供）
        blanks = sub[sub["机会信号"] == "⚡ 高需低供"]["标签值"].tolist()

        json_payload["snapshots"][cat] = {
            "top_flavor":  top1("口味"),
            "top_body":    top1("商品主体"),
            "top_price":   top1("价格带"),
            "flavors":     flavors,
            "blanks":      blanks,
        }

    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(json_payload, f, ensure_ascii=False, indent=2)
    print(f"      品类数：{len(cats)}，热度条目：{len(json_payload['heat'])}")

    print(f"\n✅ 完成！→ {OUTPUT_FILE}")
    print(f"           → {JSON_FILE}  （前端页面读取此文件）")
    print("\nSheet 说明：")
    print("  标签热度总览    所有标签维度，按品类+维度+热度排序")
    print("  口味热度        仅口味维度，采购人员最常看")
    print("  商品主体热度    仅商品主体维度")
    print("  规格价格带热度  规格数值 + 价格带")
    print("  商品宽表(下钻)  原始数据+解析字段，供具体商品查询")
    print("  热度评分说明    评分逻辑完整说明")
    print("\n词典补充提示：")
    print("  口味识别遗漏 → 在脚本顶部 FLAVOR_WORDS 里添加")
    print("  主体识别遗漏 → 在 PRODUCT_TYPES 里添加（具体词放前面）")


if __name__ == "__main__":
    main()
